"""
inspeccion_vehicular.py — v5.0
===============================
Standalone : streamlit run inspeccion_vehicular.py
Desde hub  : import inspeccion_vehicular; inspeccion_vehicular.run()

Cambios v5.0
------------
• Botón "Volver al Hub" SIEMPRE visible en todas las pantallas
• modo_falla OBLIGATORIO cuando estado = Malo/Mala (bloquea guardado)
• Mensaje claro "Se ha generado la OT XXXXXX" cuando se produce OT-CI
• URL sync: al navegar al hub actualiza query_params
"""
import streamlit as st
import pandas as pd
import os
import io
import json
from datetime import datetime
from pathlib import Path

try:
    from streamlit_drawable_canvas import st_canvas
    from PIL import Image
    CANVAS_DISPONIBLE = True
except ImportError:
    CANVAS_DISPONIBLE = False

PENDIENTES_DB = "ots_pendientes.json"
EXCEL_FILE    = "inspecciones_vehiculares.xlsx"
FLOTA_DB      = "flota_vehiculos.json"

# ═══════════════════════════════════════════════════════════════════
#  CATÁLOGO — MODOS DE FALLA POR ÍTEM DE INSPECCIÓN
# ═══════════════════════════════════════════════════════════════════
MODOS_FALLA_INSP = {
    "Frenos de servicio (Verificar fugas)": [
        "Ruido al frenar (chirrido, raspado o golpe metálico)",
        "Pedal esponjoso, blando o con recorrido excesivo",
        "Vibración al frenar",
        "Desviación lateral al frenar",
        "Fuga visible de líquido de frenos",
        "Pastillas o zapatas con desgaste visible",
        "Fuga de freno de aire (escape de aire audible)",
    ],
    "Freno de parqueo": ["Freno de parqueo no retiene el vehículo"],
    "Nivel de aceite motor": ["Nivel de aceite bajo", "Fuga visible de aceite", "Aceite con aspecto anormal"],
    "Nivel de refrigerante": [
        "Nivel bajo al revisar depósito",
        "Fuga visible de refrigerante",
        "Mangueras de refrigeración con rajaduras o fugas",
    ],
    "Nivel de liquido de frenos": ["Nivel bajo", "Fuga visible"],
    "Nivel de direccion hidraulica": ["Nivel bajo", "Fuga visible de fluido de dirección"],
    "Suspension": [
        "Ruido en suspensión al pasar por irregularidades",
        "Rebote excesivo de la carrocería",
        "Inclinación visible del vehículo hacia un lado",
        "Resorte roto o deformado",
    ],
    "Sistema de escape": [
        "Humo excesivo por el escape",
        "Ruido anormal en el escape",
        "Fuga visible de gases",
    ],
    "Tablero de Instrumentos": [
        "Testigo de alarma activo en tablero",
        "Indicador de temperatura en zona roja",
        "Testigo de batería o carga encendido",
        "Testigo de presión de aceite encendido",
    ],
    "Cinturon de seguridad": ["Cinturón no engancha, no retrae o está dañado"],
    "Plumillas / limpiaparabrisas": ["Plumillas deterioradas (rayaduras o no limpian bien)",
                                      "Plumillas no funcionan"],
    "Espejos retrovisores": ["Espejo roto, suelto o desajustado"],
    "Llantas (presion y estado)": [
        "Llanta pinchada o desinflada",
        "Desgaste irregular visible",
        "Labrado por debajo del indicador mínimo",
        "Daño visible en flanco (corte, deformación)",
        "Protuberancia (hernia) visible en llanta",
    ],
    "Luces delanteras (Altas y bajas)": ["Luces apagadas o con parpadeo",
                                          "Faro roto o con entrada de humedad"],
    "Luces traseras y stop": ["Luces apagadas o con parpadeo"],
    "Luces de reversa":      ["Luces de reversa no funcionan"],
    "Luces direccionales":   ["Luces direccionales no funcionan o parpadeo anormal"],
    "Funcionamiento de puertas": [
        "Puertas no abren o no cierran correctamente",
        "Mecanismo de apertura/cierre con daño visible",
    ],
    "Pasamanos":           ["Pasamanos suelto, roto o con fijación débil"],
    "Silleteria":          ["Silletería rota, rasgada o con soporte deteriorado"],
    "Pito / bocina":       ["Bocina sin funcionamiento"],
    "Equipo de carretera": ["Equipo de carretera incompleto"],
    "Ventanas y panoramico": ["Fisura o rotura en parabrisas o vidrios"],
    "Plataforma Movilidad Reducida": ["Plataforma inoperante o con daño visible"],
}

# ═══════════════════════════════════════════════════════════════════
#  ÍTEMS AGRUPADOS POR SISTEMA
# ═══════════════════════════════════════════════════════════════════
GRUPOS_INSPECCION = {
    "Frenos": [
        ("Frenos de servicio (Verificar fugas)", True,  "M"),
        ("Freno de parqueo",                     True,  "M"),
    ],
    "Sistema Chasis": [
        ("Nivel de aceite motor",                True,  "M"),
        ("Nivel de refrigerante",                True,  "M"),
        ("Nivel de liquido de frenos",           True,  "M"),
        ("Nivel de direccion hidraulica",        False, "M"),
        ("Suspension",                           True,  "F"),
        ("Sistema de escape",                    True,  "M"),
    ],
    "Habitaculo": [
        ("Pasamanos",                            False, "M"),
        ("Silleteria",                           False, "F"),
        ("Tablero de Instrumentos",              True,  "M"),
        ("Cinturon de seguridad",                True,  "M"),
        ("Pito / bocina",                        False, "M"),
        ("Equipo de carretera",                  False, "M"),
    ],
    "Carroceria Externa": [
        ("Ventanas y panoramico",                False, "F"),
        ("Plumillas / limpiaparabrisas",         True,  "F"),
        ("Espejos retrovisores",                 True,  "M"),
        ("Llantas (presion y estado)",           True,  "F"),
        ("Luces delanteras (Altas y bajas)",     False, "F"),
        ("Luces traseras y stop",                False, "F"),
        ("Luces de reversa",                     False, "F"),
        ("Luces direccionales",                  False, "F"),
        ("Funcionamiento de puertas",            True,  "M"),
        ("Plataforma Movilidad Reducida",        False, "F"),
    ],
}
ITEMS_INSPECCION = [(i, c, g) for grp in GRUPOS_INSPECCION.values() for i, c, g in grp]

# ═══════════════════════════════════════════════════════════════════
#  MAPEO ÍTEM → SISTEMA ESTÁNDAR (para OTs generadas)
#  Los ítems de la lista de chequeo se agrupan en los mismos sistemas
#  que usa el módulo de Reporte de Fallas, para garantizar coherencia
#  analítica en Power BI. Los grupos VISUALES (GRUPOS_INSPECCION) no
#  cambian — solo el sistema que se inscribe en la OT.
# ═══════════════════════════════════════════════════════════════════
ITEM_TO_SISTEMA = {
    # — Motor —
    "Nivel de aceite motor":            "Motor",
    "Sistema de escape":                "Motor",
    # — Frenos —
    "Frenos de servicio (Verificar fugas)": "Frenos",
    "Freno de parqueo":                 "Frenos",
    "Nivel de liquido de frenos":       "Frenos",
    # — Dirección —
    "Nivel de direccion hidraulica":    "Direccion",
    # — Suspensión —
    "Suspension":                       "Suspension",
    # — Llantas —
    "Llantas (presion y estado)":       "Llantas",
    # — Refrigeración —
    "Nivel de refrigerante":            "Refrigeracion",
    # — Eléctrico —
    "Tablero de Instrumentos":          "Electrico",
    "Luces delanteras (Altas y bajas)": "Electrico",
    "Luces traseras y stop":            "Electrico",
    "Luces de reversa":                 "Electrico",
    "Luces direccionales":              "Electrico",
    # — Carrocería Externa —
    "Ventanas y panoramico":            "Carroceria Externa",
    "Plumillas / limpiaparabrisas":     "Carroceria Externa",
    "Espejos retrovisores":             "Carroceria Externa",
    "Funcionamiento de puertas":        "Carroceria Externa",
    "Plataforma Movilidad Reducida":    "Carroceria Externa",
    # — Habitáculo —
    "Pasamanos":                        "Habitaculo",
    "Silleteria":                       "Habitaculo",
    "Cinturon de seguridad":            "Habitaculo",
    "Pito / bocina":                    "Habitaculo",
    "Equipo de carretera":              "Habitaculo",
}


def _agrupar_novedades_por_sistema(session_inspeccion):
    """
    Recorre los items con novedad y los agrupa por sistema estándar.
    Retorna: dict {sistema: [(item, estado, modo_falla, es_critico), ...]}
    """
    from collections import defaultdict
    agrupado = defaultdict(list)
    for item_name, es_critico, _ in ITEMS_INSPECCION:
        datos = session_inspeccion.get(item_name, {})
        estado = datos.get("estado", "")
        if _es_bueno(estado):
            continue
        sistema = ITEM_TO_SISTEMA.get(item_name, "Carroceria Externa")
        modo = datos.get("modo_falla", "") or ""
        agrupado[sistema].append((item_name, estado, modo, es_critico))
    return dict(agrupado)


def _opciones(g):     return ["Bueno", "Regular", "Malo"] if g == "M" else ["Buena", "Regular", "Mala"]
def _estado_bueno(g): return "Bueno" if g == "M" else "Buena"
def _es_bueno(e):     return e in ("Buena", "Bueno")
def _es_malo(e):      return e in ("Malo", "Mala")

def _calc_sem(estado, critico):
    if _es_bueno(estado):            return "ok"
    if _es_malo(estado) and critico: return "critical"
    return "alert"

ESTADO_INICIAL = {i: _estado_bueno(g) for i, _, g in ITEMS_INSPECCION}
DOCUMENTOS = ["SOAT", "Revision Tecnico-Mecanica", "Tarjeta de Operacion", "Licencia de Conduccion"]


# ═══════════════════════════════════════════════════════════════════
#  PERSISTENCIA
# ═══════════════════════════════════════════════════════════════════
def _save_excel(row_data):
    try:
        df_new = pd.DataFrame([row_data])
        if os.path.exists(EXCEL_FILE):
            try:
                df_old = pd.read_excel(EXCEL_FILE, engine="openpyxl")
                df_all = pd.concat([df_old, df_new], ignore_index=True)
            except PermissionError:
                return False, "El archivo Excel está abierto. Ciérrelo e intente de nuevo."
        else:
            df_all = df_new
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as w:
            df_all.to_excel(w, index=False, sheet_name="Inspecciones")
            ws = w.sheets["Inspecciones"]
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            hf = PatternFill("solid", fgColor="0A1628")
            for cell in ws[1]:
                cell.fill = hf
                cell.font = Font(color="FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            rf = PatternFill("solid", fgColor="FFCCCC")
            yf = PatternFill("solid", fgColor="FFF3CC")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        v = cell.value.strip().lower()
                        if v.startswith("malo") or v.startswith("mala"):
                            cell.fill = rf
                        elif v.startswith("regular"):
                            cell.fill = yf
            for col in ws.columns:
                ml = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 45)
        return True, "OK"
    except PermissionError:
        return False, "El archivo Excel está abierto."
    except Exception as e:
        return False, str(e)


def _carpeta_vehiculo(num):
    p = Path("archivos") / f"{datetime.now().strftime('%Y-%m-%d')}_{num.upper().replace(' ', '_')}"
    p.mkdir(parents=True, exist_ok=True)
    return p

def _save_image(f, num):
    fp = _carpeta_vehiculo(num) / f"{num.upper()}_{datetime.now().strftime('%H%M%S')}.{f.name.split('.')[-1]}"
    fp.write_bytes(f.read())
    return str(fp)

def _save_firma_bytes(png_bytes, num):
    fp = _carpeta_vehiculo(num) / f"firma_{num.upper()}_{datetime.now().strftime('%H%M%S')}.png"
    fp.write_bytes(png_bytes)
    return str(fp)


def _generar_ot_ci(num, sistema=""):
    """Genera radicado OT-CI con abreviatura del sistema para trazabilidad.
       Formato: OT-CI-<SIST>-<VEH>-<YYMMDD>-<HHMM>"""
    veh = num.strip().upper().replace(" ", "")[:6]
    # Abreviaturas de 3 letras por sistema
    sist_abrev = {
        "Motor":              "MOT",
        "Frenos":             "FRE",
        "Direccion":          "DIR",
        "Suspension":         "SUS",
        "Llantas":            "LLA",
        "Refrigeracion":      "REF",
        "Electrico":          "ELE",
        "Carroceria Externa": "CAR",
        "Habitaculo":         "HAB",
        "Transmision":        "TRA",
    }.get(sistema, "VEH")
    ts = datetime.now()
    return f"OT-CI-{sist_abrev}-{veh}-{ts.strftime('%y%m%d')}-{ts.strftime('%H%M')}"

def _registrar_ot_ci(ot, datos):
    db = {}
    if os.path.exists(PENDIENTES_DB):
        try:
            with open(PENDIENTES_DB, encoding="utf-8") as f:
                db = json.load(f)
        except Exception:
            pass
    db[ot] = datos
    with open(PENDIENTES_DB, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)


def _cargar_lista_vehiculos():
    """Retorna lista ordenada de IDs de vehículos en la flota, o lista vacía si no hay."""
    if not os.path.exists(FLOTA_DB):
        return []
    try:
        with open(FLOTA_DB, encoding="utf-8") as f:
            flota = json.load(f)
        return sorted(flota.keys())
    except Exception:
        return []


def _info_vehiculo(veh_id):
    """Retorna dict con info del vehículo o None si no existe."""
    if not veh_id or not os.path.exists(FLOTA_DB):
        return None
    try:
        with open(FLOTA_DB, encoding="utf-8") as f:
            flota = json.load(f)
        return flota.get(veh_id.upper())
    except Exception:
        return None


def _actualizar_flota_desde_inspeccion(numero_interno, km_inspeccion):
    """
    Auto-actualiza el kilometraje en flota_vehiculos.json al guardar una inspección.
    Comportamiento:
    - Si el vehículo NO existe en flota_vehiculos.json, se crea automáticamente
      (autoregistro silencioso) con el km de la inspección como valor inicial.
    - Si el vehículo existe y km_inspeccion > km_actual, se actualiza.
    - Si km_inspeccion <= km_actual, no se modifica (los buses no rebobinan).
    Retorna: (acción, km_anterior) — acción ∈ {"creado", "actualizado", "sin_cambio"}
    """
    if not numero_interno or km_inspeccion <= 0:
        return ("sin_cambio", 0)

    veh_key = numero_interno.strip().upper()
    flota = {}
    if os.path.exists(FLOTA_DB):
        try:
            with open(FLOTA_DB, encoding="utf-8") as f:
                flota = json.load(f)
        except Exception:
            flota = {}

    fecha_hoy = datetime.now().strftime("%Y-%m-%d")

    if veh_key not in flota:
        # Auto-registro silencioso de vehículo nuevo
        flota[veh_key] = {
            "km_actual":         km_inspeccion,
            "km_actualizado_en": fecha_hoy,
            "rutinas_ultimas":   {},
            "auto_registrado":   True,
        }
        accion = "creado"
        km_anterior = 0
    else:
        km_anterior = flota[veh_key].get("km_actual", 0)
        if km_inspeccion > km_anterior:
            flota[veh_key]["km_actual"]         = km_inspeccion
            flota[veh_key]["km_actualizado_en"] = fecha_hoy
            accion = "actualizado"
        else:
            accion = "sin_cambio"

    try:
        with open(FLOTA_DB, "w", encoding="utf-8") as f:
            json.dump(flota, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    return (accion, km_anterior)


# ═══════════════════════════════════════════════════════════════════
#  NAVEGACIÓN AL HUB
# ═══════════════════════════════════════════════════════════════════
def _volver_al_hub():
    """Limpia estado del módulo y regresa al hub (persistente en URL)."""
    for k in ["submitted", "finalizado", "saved_data", "inspeccion", "docs",
              "firma_png", "contador", "despacho", "ot_ci_generada", "insp_errores_modo"]:
        st.session_state.pop(k, None)
    st.session_state["modulo"] = "hub"
    try:
        st.query_params.clear()
    except Exception:
        pass
    st.rerun()


# ═══════════════════════════════════════════════════════════════════
#  CSS
# ═══════════════════════════════════════════════════════════════════
CSS = """<style>
@import url('https://fonts.googleapis.com/css2?family=Exo+2:wght@700;800&family=Inter:wght@300;400;500;600&display=swap');
:root{--primary:#0A84FF;--primary-dk:#005FCC;--bg:#0D1117;--surface:#161B22;--surface2:#21262D;
      --border:#30363D;--text:#E6EDF3;--muted:#8B949E;--green:#3FB950;--yellow:#D29922;
      --red:#F85149;--orange:#FF6B00;--radius:12px;}
html,body,[data-testid="stAppViewContainer"]{background:var(--bg)!important;color:var(--text)!important;font-family:'Inter',sans-serif}
.back-to-hub{display:inline-flex;align-items:center;gap:6px;background:var(--surface2);border:1px solid var(--border);border-radius:20px;padding:5px 14px;font-size:.78rem;color:var(--muted);margin-bottom:.8rem;cursor:pointer;transition:all .15s}
.back-to-hub:hover{border-color:var(--text);color:var(--text)}
.app-header{background:linear-gradient(135deg,#0D1117,#161B22,#0A1628);border-bottom:1px solid var(--border);padding:1.3rem 2rem;margin:0 -1rem 2rem -1rem;display:flex;align-items:center;gap:1rem}
.app-header h1{font-family:'Exo 2',sans-serif;font-weight:800;font-size:1.5rem;color:var(--text);margin:0}
.app-header .sub{font-size:.78rem;color:var(--muted);margin:0}
.badge-pill{background:var(--primary);color:white;padding:.2rem .8rem;border-radius:20px;font-size:.7rem;font-weight:600;text-transform:uppercase}
.card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:1.5rem;margin-bottom:1.2rem}
.card-title{font-family:'Exo 2',sans-serif;font-size:.95rem;font-weight:700;color:var(--primary);margin-bottom:1rem;text-transform:uppercase;letter-spacing:.05em}
.group-hdr{background:linear-gradient(90deg,rgba(10,132,255,.15),transparent);border-left:3px solid var(--primary);border-radius:0 8px 8px 0;padding:.6rem 1rem;margin:1.2rem 0 .8rem;font-family:'Exo 2',sans-serif;font-weight:700;font-size:.9rem;color:var(--text);text-transform:uppercase;letter-spacing:.08em}
.item-row{background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:.7rem 1rem;margin-bottom:.4rem}
.item-row.err-row{border-color:var(--red);background:rgba(248,81,73,.05)}
.item-name{font-weight:600;font-size:.9rem;color:var(--text)}
.tag-crit{font-size:.7rem;color:var(--red);font-weight:600;margin-top:2px}
.tag-norm{font-size:.7rem;color:var(--muted);margin-top:2px}
.tag-err{font-size:.7rem;color:var(--red);font-weight:700;margin-top:2px}
.sem-bar{display:flex;gap:12px;margin:12px 0;flex-wrap:wrap}
.sem-box{display:flex;align-items:center;gap:6px;background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:6px 14px;font-size:.85rem;font-weight:600}
.sem-ok{border-color:#3FB950;color:#3FB950}.sem-alert{border-color:#D29922;color:#D29922}.sem-crit{border-color:#F85149;color:#F85149}
.desp-si{background:rgba(63,185,80,.07);border:1px solid #3FB950;border-radius:10px;padding:1rem 1.2rem;margin:.5rem 0}
.desp-no{background:rgba(248,81,73,.07);border:1px solid #F85149;border-radius:10px;padding:1rem 1.2rem;margin:.5rem 0}
.ot-box-generated{background:linear-gradient(135deg,#1A1000,#0D1117);border:2px solid #FF6B00;border-radius:12px;padding:1.2rem 1.5rem;margin:1rem 0;box-shadow:0 0 24px rgba(255,107,0,.2)}
.ot-box-label{font-size:.72rem;color:#8B949E;text-transform:uppercase;letter-spacing:.1em;font-weight:600}
.ot-box-number{font-family:'Exo 2',sans-serif;font-size:1.5rem;font-weight:800;color:#FF6B00;margin:.3rem 0}
.ot-box-info{font-size:.8rem;color:#8B949E;margin-top:.3rem}
.success-screen{text-align:center;padding:2.5rem 2rem;background:linear-gradient(135deg,#0D1F3C,#0D1117);border:1px solid var(--primary);border-radius:var(--radius);margin:1.5rem 0}
.success-screen h2{font-family:'Exo 2',sans-serif;font-size:1.8rem;color:var(--text);margin:.4rem 0}
.final-screen{text-align:center;padding:3rem 2rem;background:linear-gradient(135deg,#1A0A0A,#0D1117);border:1px solid var(--red);border-radius:var(--radius);margin:2rem 0}
.error-box{background:rgba(248,81,73,.08);border:1px solid var(--red);border-radius:10px;padding:.8rem 1.2rem;margin:.8rem 0;font-size:.85rem;color:#F85149}
[data-testid="stTextInput"] input,[data-testid="stTextArea"] textarea{background:var(--surface2)!important;border:1px solid var(--border)!important;color:var(--text)!important;border-radius:8px!important;font-family:'Inter',sans-serif!important}
[data-testid="stTextInput"] input:focus{border-color:var(--primary)!important;box-shadow:0 0 0 2px rgba(10,132,255,.2)!important}
[data-testid="stSelectbox"]>div>div{background:var(--surface2)!important;border:1px solid var(--border)!important;border-radius:8px!important}
[data-testid="stRadio"]>div{gap:.4rem!important}
[data-testid="stRadio"] label{background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:.35rem .9rem!important;cursor:pointer;font-size:.85rem!important}
[data-testid="stRadio"] label:has(input:checked){border-color:var(--primary)!important;background:rgba(10,132,255,.1)!important}
[data-testid="stButton"] button{background:var(--primary)!important;color:white!important;border:none!important;border-radius:8px!important;font-family:'Exo 2',sans-serif!important;font-weight:700!important;font-size:.9rem!important;padding:.55rem 1.2rem!important;text-transform:uppercase!important;letter-spacing:.05em!important}
[data-testid="stButton"] button:hover{background:var(--primary-dk)!important;transform:translateY(-1px);box-shadow:0 4px 16px rgba(10,132,255,.3)!important}
[data-testid="stButton"] button:disabled{background:var(--surface2)!important;color:var(--muted)!important;transform:none!important;box-shadow:none!important}
.btn-green button{background:#2EA043!important}.btn-green button:hover{background:#3FB950!important}
.btn-red button{background:#6E1A19!important}.btn-red button:hover{background:var(--red)!important}
.btn-back button{background:var(--surface2)!important;color:var(--muted)!important;border:1px solid var(--border)!important;text-transform:none!important;font-size:.82rem!important;font-weight:500!important;padding:.4rem 1rem!important;letter-spacing:normal!important}
.btn-back button:hover{background:var(--border)!important;color:var(--text)!important;transform:none!important;box-shadow:none!important}
.firma-ok{border:1px solid var(--green);border-radius:8px;padding:.5rem 1rem;background:rgba(63,185,80,.07);color:var(--green);font-size:.85rem;margin-top:.4rem}
.firma-req{border:1px solid var(--red);border-radius:8px;padding:.5rem 1rem;background:rgba(248,81,73,.07);color:var(--red);font-size:.85rem;margin-top:.4rem}
.prog-wrap{margin:1rem 0}.prog-lbl{font-size:.8rem;color:var(--muted);margin-bottom:.3rem}
.prog-bar{height:6px;background:var(--surface2);border-radius:3px;overflow:hidden}
.prog-fill{height:100%;background:linear-gradient(90deg,var(--primary),#00C9FF);border-radius:3px}
.sec-div{border:none;border-top:1px solid var(--border);margin:.8rem 0}
#MainMenu,footer,header{visibility:hidden}.block-container{padding-top:0!important}
@media(max-width:768px){.app-header h1{font-size:1.2rem}.card{padding:1rem}}
</style>"""


# ═══════════════════════════════════════════════════════════════════
#  FUNCIÓN PRINCIPAL
# ═══════════════════════════════════════════════════════════════════
def _init_state():
    defaults = {
        "submitted":         False,
        "finalizado":        False,
        "saved_data":        None,
        "contador":          0,
        "firma_png":         None,
        "despacho":          "Si",
        "ot_ci_generada":    None,
        "inspeccion":        {i: {"estado": ESTADO_INICIAL[i], "obs": "", "modo_falla": ""}
                              for i, _, _ in ITEMS_INSPECCION},
        "docs":              {d: "Vigente" for d in DOCUMENTOS},
        "insp_errores_modo": [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _reset_form():
    for k in ["submitted", "saved_data", "firma_png", "despacho",
              "ot_ci_generada", "finalizado", "insp_errores_modo"]:
        st.session_state.pop(k, None)
    st.session_state["inspeccion"] = {
        i: {"estado": ESTADO_INICIAL[i], "obs": "", "modo_falla": ""}
        for i, _, _ in ITEMS_INSPECCION
    }
    st.session_state["docs"] = {d: "Vigente" for d in DOCUMENTOS}


def _render_back_button(key_suffix=""):
    """Botón Volver al Hub — visible en todas las pantallas cuando se ejecuta desde el hub."""
    if st.session_state.get("modulo") == "inspeccion":
        st.markdown('<div class="btn-back">', unsafe_allow_html=True)
        if st.button("← Volver al Hub", key=f"btn_back_hub_{key_suffix}"):
            _volver_al_hub()
        st.markdown('</div>', unsafe_allow_html=True)


def _header(v="v5.0"):
    st.markdown(f"""
    <div class="app-header">
      <div style="font-size:2rem">🔍</div>
      <div><h1>Inspección Preoperacional</h1>
           <p class="sub">Vehículos de Transporte — Control Técnico Diario</p></div>
      <div style="margin-left:auto"><span class="badge-pill">{v}</span></div>
    </div>""", unsafe_allow_html=True)


def run():
    _init_state()
    st.markdown(CSS, unsafe_allow_html=True)

    # ── PANTALLA: FINALIZADO ─────────────────────────────
    if st.session_state.get("finalizado"):
        _render_back_button("final")
        _header()
        n = st.session_state.get("contador", 0)
        st.markdown(f"""<div class="final-screen">
          <div style="font-size:4rem">🏁</div>
          <h2 style="color:var(--red);font-family:'Exo 2',sans-serif">Sesión finalizada</h2>
          <p style="color:#8B949E">Se registraron {n} inspección(es) en esta sesión.</p>
        </div>""", unsafe_allow_html=True)
        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("🔄 Nueva sesión", use_container_width=True):
                _reset_form()
                st.rerun()
        with col_b:
            st.markdown('<div class="btn-back">', unsafe_allow_html=True)
            if st.button("🏠 Ir al Hub", use_container_width=True, key="btn_hub_final"):
                _volver_al_hub()
            st.markdown('</div>', unsafe_allow_html=True)
        return

    # ── PANTALLA: CONFIRMACIÓN POST-GUARDADO ─────────────
    if st.session_state.get("submitted"):
        _render_back_button("subm")
        data     = st.session_state.get("saved_data", {})
        ot_ci    = st.session_state.get("ot_ci_generada")
        despacho = st.session_state.get("despacho", "Si")

        _header()
        st.markdown("""<div class="success-screen">
          <div style="font-size:4rem">✅</div>
          <h2>Inspección Preoperacional Ejecutada</h2>
          <p style="color:#8B949E;margin:.4rem 0 0">Registro almacenado correctamente.</p>
        </div>""", unsafe_allow_html=True)

        if data:
            c1, c2, c3 = st.columns(3)
            c1.metric("Prealistador", data.get("Prealistador", "—"))
            c2.metric("N° Interno",   data.get("Numero_Interno", "—"))
            c3.metric("Fecha/Hora",   data.get("Fecha_Hora", "—"))
            # Bugfix: el valor guardado en `data` puede ser "Malo | modo | obs",
            # hay que extraer solo el estado (primer token antes del separador) antes de evaluarlo.
            def _estado_de(item_name):
                raw = str(data.get(item_name, ""))
                return raw.split("|")[0].strip() if raw else ""

            criticos = sum(1 for n, c, _ in ITEMS_INSPECCION
                           if _calc_sem(_estado_de(n), c) == "critical")
            alertas  = sum(1 for n, c, _ in ITEMS_INSPECCION
                           if _calc_sem(_estado_de(n), c) == "alert")
            oks      = len(ITEMS_INSPECCION) - criticos - alertas
            ca, cb, cc = st.columns(3)
            ca.metric("🟢 OK", oks)
            cb.metric("🟡 Alertas", alertas)
            cc.metric("🔴 Críticos", criticos)

        # ── Mensaje de actualización de flota (silencioso) ──
        accion_flota = st.session_state.get("flota_accion")
        km_anterior  = st.session_state.get("flota_km_ant", 0)
        veh_id_msg   = data.get("Numero_Interno", "") if data else ""
        km_actual_msg = data.get("Kilometraje", "") if data else ""

        if accion_flota == "creado":
            st.markdown(f"""<div style="background:rgba(124,77,255,.08);border-left:3px solid #7C4DFF;border-radius:6px;padding:.6rem 1rem;margin:.8rem 0;font-size:.82rem;color:#B47AFF">
              🚛 <strong>Vehículo {veh_id_msg} registrado automáticamente</strong> en la flota con km inicial:
              <strong>{int(km_actual_msg):,} km</strong>. Las alarmas de mantenimiento preventivo
              se generarán a partir de este registro.
            </div>""".replace(",", "."), unsafe_allow_html=True)
        elif accion_flota == "actualizado" and km_anterior > 0:
            delta_km = int(km_actual_msg) - km_anterior
            st.markdown(f"""<div style="background:rgba(43,126,255,.06);border-left:3px solid #2B7EFF;border-radius:6px;padding:.6rem 1rem;margin:.8rem 0;font-size:.82rem;color:#60A5FF">
              📊 <strong>Kilometraje actualizado en flota</strong>:
              {km_anterior:,} km → <strong>{int(km_actual_msg):,} km</strong>
              <span style="color:#8B949E">(+{delta_km:,} km desde último registro)</span>
            </div>""".replace(",", "."), unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── OTs GENERADAS (una por sistema con novedad) ──
        ots_generadas_str = data.get("OT_Correctiva_Generada", "") if data else ""
        lista_ots = [x.strip() for x in ots_generadas_str.split(" | ") if x.strip()] if ots_generadas_str else []

        if lista_ots:
            n_ots = len(lista_ots)
            titulo = f"✅ Se generó {n_ots} Orden de Trabajo Correctiva" if n_ots == 1 \
                     else f"✅ Se generaron {n_ots} Órdenes de Trabajo Correctivas"
            subtitulo = "(una por cada sistema con novedad — coherente con el módulo de Reporte de Fallas)" \
                        if n_ots > 1 else ""

            ot_numbers_html = "<br>".join(
                f'<span style="font-family:Exo 2,sans-serif;font-size:1.15rem;font-weight:800;color:#FF6B00">{ot}</span>'
                for ot in lista_ots
            )
            st.markdown(f"""<div class="ot-box-generated">
              <div class="ot-box-label">{titulo}</div>
              <div style="margin:.5rem 0">{ot_numbers_html}</div>
              <div class="ot-box-info">
                {subtitulo}<br>
                Acceda al módulo <strong style="color:#E6EDF3">Cierre de OT</strong> para que el mecánico
                complete las intervenciones técnicas.
              </div>
            </div>""", unsafe_allow_html=True)
            if n_ots > 1:
                st.info(f"📋 Se crearon {n_ots} OTs independientes — cada una agrupando los ítems de su respectivo sistema. "
                        f"Esto permite trazabilidad y análisis por sistema en el dashboard.")
            else:
                st.info("📋 La OT quedó registrada como pendiente. Puede ir directamente al módulo de cierre con el botón de abajo.")

        if despacho == "No":
            st.markdown('<div class="desp-no"><strong style="color:#F85149">🚫 Despacho NO autorizado</strong> — '
                        '<span style="color:#8B949E;font-size:.85rem">Vehículo retenido por novedades.</span></div>',
                        unsafe_allow_html=True)
        else:
            st.markdown('<div class="desp-si"><strong style="color:#3FB950">✅ Despacho AUTORIZADO</strong> — '
                        '<span style="color:#8B949E;font-size:.85rem">Vehículo sale a operación.</span></div>',
                        unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        col_s, col_f = st.columns(2)
        with col_s:
            st.markdown('<div class="btn-green">', unsafe_allow_html=True)
            if st.button("🚛 Siguiente Vehículo", use_container_width=True, key="btn_sig"):
                _reset_form()
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        with col_f:
            st.markdown('<div class="btn-red">', unsafe_allow_html=True)
            if st.button("🏁 Finalizar Proceso", use_container_width=True, key="btn_fin"):
                _reset_form()
                st.session_state["finalizado"] = True
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        if lista_ots and st.session_state.get("modulo") == "inspeccion":
            st.markdown("<br>", unsafe_allow_html=True)
            if len(lista_ots) == 1:
                btn_label = f"🔧 Ir al Cierre de OT  ({lista_ots[0]})"
            else:
                btn_label = f"🔧 Ir al módulo de Cierre de OT  ({len(lista_ots)} OTs pendientes)"
            if st.button(btn_label, use_container_width=True, key="btn_goto_ot"):
                st.session_state["modulo"]        = "cierre_ot"
                st.session_state["fase"]          = "2"
                # Si solo hay una, precargarla directamente; si son varias, que elija
                st.session_state["ot_actual"]     = lista_ots[0] if len(lista_ots) == 1 else None
                st.session_state["intervenciones"] = [{"desc": "", "cantidad": 1, "costo": 0.0}]
                try:
                    st.query_params["mod"] = "cierre_ot"
                except Exception:
                    pass
                st.rerun()
        return

    # ════════════════════════════════════════════════════
    #  FORMULARIO PRINCIPAL
    # ════════════════════════════════════════════════════
    _render_back_button("form")
    _header()
    st.markdown(f"<p style='color:#8B949E;font-size:.8rem;margin-bottom:1.5rem'>📅 {datetime.now().strftime('%A, %d de %B de %Y — %H:%M')}</p>",
                unsafe_allow_html=True)

    # ── Datos básicos ──────────────────────────────
    st.markdown('<div class="card"><div class="card-title">Datos del Prealistamiento</div>',
                unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        prealistador = st.text_input("Nombre del Prealistador *",
                                      placeholder="Ej: Carlos Ramírez", key="prealistador")
    with c2:
        # Si hay flota registrada, mostrar selectbox; si no, text_input libre
        flota_ids = _cargar_lista_vehiculos()
        if flota_ids:
            opciones = ["-- Seleccione --"] + flota_ids + ["+ Otro vehículo (escribir)"]
            sel = st.selectbox("Número Interno del Vehículo *",
                                 options=opciones, key="numero_interno_sel")
            if sel == "+ Otro vehículo (escribir)":
                numero_interno = st.text_input(" ", placeholder="Ej: 0042",
                                                 key="numero_interno_libre",
                                                 label_visibility="collapsed")
            elif sel == "-- Seleccione --":
                numero_interno = ""
            else:
                numero_interno = sel
                # Mostrar info del vehículo seleccionado
                info = _info_vehiculo(sel)
                if info:
                    marca = info.get("marca", "")
                    modelo = info.get("modelo", "")
                    if marca or modelo:
                        st.markdown(f"<div style='font-size:.78rem;color:var(--muted);margin-top:-.3rem'>"
                                    f"🚛 {marca} {modelo}</div>", unsafe_allow_html=True)
        else:
            numero_interno = st.text_input("Número Interno del Vehículo *",
                                            placeholder="Ej: 0042", key="numero_interno")
            st.caption("ℹ️ El primer vehículo registrado quedará automáticamente en la flota.")
    c3, c4 = st.columns(2)
    with c3:
        conductor = st.text_input("Conductor (opcional)",
                                   placeholder="Ej: Pedro Gómez", key="conductor")
    with c4:
        # Si hay vehículo seleccionado de flota, mostrar último km registrado como referencia
        info_veh = _info_vehiculo(numero_interno) if numero_interno else None
        ultimo_km_flota = info_veh.get("km_actual", 0) if info_veh else 0

        km_int = st.number_input("Kilometraje *", min_value=0, step=100, value=0,
                                  key="km", format="%d")
        if km_int > 0:
            st.markdown(f"<div style='font-size:.85rem;color:var(--primary);font-weight:600;margin-top:-.4rem'>"
                        f"{km_int:,} km".replace(",", ".") + "</div>", unsafe_allow_html=True)
        elif ultimo_km_flota > 0:
            st.markdown(f"<div style='font-size:.75rem;color:var(--muted);margin-top:-.4rem'>"
                        f"📊 Último km registrado: <strong>{ultimo_km_flota:,}</strong>"
                        .replace(",", ".") + "</div>", unsafe_allow_html=True)
        km = str(km_int) if km_int > 0 else ""
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Lista de chequeo ───────────────────────────
    st.markdown('<div class="card"><div class="card-title">Lista de Chequeo Técnico</div>',
                unsafe_allow_html=True)
    st.markdown("""<div style="display:flex;gap:16px;margin-bottom:12px;flex-wrap:wrap;font-size:.8rem;color:#8B949E">
      <span>🟢 OK = Bueno/Buena</span><span>🟡 ALERTA = Regular</span>
      <span>🔴 CRÍTICO = Malo en sistema crítico</span>
    </div>
    <div style="background:rgba(248,81,73,.05);border:1px solid rgba(248,81,73,.2);border-radius:8px;padding:.6rem 1rem;margin-bottom:1rem;font-size:.78rem;color:#F85149">
      ⚠️ <strong>Si marca un ítem como Malo/Mala, es obligatorio seleccionar el modo de falla</strong> (para alimentar la OT correctiva).
    </div>""", unsafe_allow_html=True)

    errores_modo_actuales = st.session_state.get("insp_errores_modo", [])

    for grupo_nombre, items in GRUPOS_INSPECCION.items():
        st.markdown(f'<div class="group-hdr">{grupo_nombre}</div>', unsafe_allow_html=True)
        for item_name, es_critico, genero in items:
            estado_actual = st.session_state["inspeccion"][item_name]["estado"]
            opciones = _opciones(genero)
            if estado_actual not in opciones:
                estado_actual = opciones[0]
                st.session_state["inspeccion"][item_name]["estado"] = estado_actual

            tiene_error = item_name in errores_modo_actuales
            row_class = "item-row err-row" if tiene_error else "item-row"
            if tiene_error:
                tag = '<span class="tag-err">⚠️ Seleccione modo de falla</span>'
            elif es_critico:
                tag = '<span class="tag-crit">Sistema Crítico</span>'
            else:
                tag = '<span class="tag-norm">No crítico</span>'

            st.markdown(f'<div class="{row_class}"><div class="item-name">{item_name}</div>{tag}</div>',
                        unsafe_allow_html=True)

            ca, cb, cc = st.columns([1, 1.5, 1.5])
            with ca:
                nuevo = st.radio(f"_e_{item_name}", opciones,
                                  index=opciones.index(estado_actual),
                                  key=f"radio_{item_name}",
                                  label_visibility="collapsed", horizontal=True)
                st.session_state["inspeccion"][item_name]["estado"] = nuevo

            with cb:
                modos = MODOS_FALLA_INSP.get(item_name, [])
                if modos and not _es_bueno(nuevo):
                    es_obligatorio = _es_malo(nuevo)
                    modo_actual = st.session_state["inspeccion"][item_name].get("modo_falla", "")
                    placeholder = "-- OBLIGATORIO: modo de falla --" if es_obligatorio else "-- Modo de falla --"
                    opts = [placeholder] + modos
                    idx = modos.index(modo_actual) + 1 if modo_actual in modos else 0
                    modo_sel = st.selectbox(f"_mf_{item_name}", options=opts,
                                             index=idx, key=f"mf_{item_name}",
                                             label_visibility="collapsed")
                    st.session_state["inspeccion"][item_name]["modo_falla"] = (
                        "" if modo_sel.startswith("--") else modo_sel
                    )
                else:
                    st.session_state["inspeccion"][item_name]["modo_falla"] = ""

            with cc:
                ob = st.text_input(f"_o_{item_name}",
                                    value=st.session_state["inspeccion"][item_name]["obs"],
                                    placeholder="Observación adicional...",
                                    key=f"obs_{item_name}", label_visibility="collapsed")
                st.session_state["inspeccion"][item_name]["obs"] = ob

            st.markdown("<hr class='sec-div'>", unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    # ── Resumen semáforo ───────────────────────────
    oks   = sum(1 for n, c, _ in ITEMS_INSPECCION
                if _calc_sem(st.session_state["inspeccion"][n]["estado"], c) == "ok")
    alts  = sum(1 for n, c, _ in ITEMS_INSPECCION
                if _calc_sem(st.session_state["inspeccion"][n]["estado"], c) == "alert")
    crits = sum(1 for n, c, _ in ITEMS_INSPECCION
                if _calc_sem(st.session_state["inspeccion"][n]["estado"], c) == "critical")
    st.markdown(f"""<div class="sem-bar">
      <div class="sem-box sem-ok">🟢 OK &nbsp;<strong>{oks}</strong></div>
      <div class="sem-box sem-alert">🟡 Alertas &nbsp;<strong>{alts}</strong></div>
      <div class="sem-box sem-crit">🔴 Críticos &nbsp;<strong>{crits}</strong></div>
    </div>""", unsafe_allow_html=True)

    # ── Documentación ──────────────────────────────
    st.markdown('<div class="card"><div class="card-title">Estado de Documentación</div>',
                unsafe_allow_html=True)
    cols_doc = st.columns(2)
    for i, doc in enumerate(DOCUMENTOS):
        with cols_doc[i % 2]:
            ed = st.radio(doc, ["Vigente", "Vencido"],
                           index=0 if st.session_state["docs"][doc] == "Vigente" else 1,
                           key=f"doc_{doc}", horizontal=True)
            st.session_state["docs"][doc] = ed
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Evidencia fotográfica ──────────────────────
    st.markdown('<div class="card"><div class="card-title">Evidencia Fotográfica</div>',
                unsafe_allow_html=True)
    uploaded_files = st.file_uploader("Cargar imágenes",
                                       type=["jpg", "jpeg", "png", "webp"],
                                       accept_multiple_files=True, key="fotos")
    if uploaded_files:
        ci_cols = st.columns(min(len(uploaded_files), 3))
        for i, f in enumerate(uploaded_files):
            with ci_cols[i % 3]:
                st.image(f, caption=f.name, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Firma digital ──────────────────────────────
    st.markdown('<div class="card"><div class="card-title">Firma Digital del Prealistador *</div>',
                unsafe_allow_html=True)
    if CANVAS_DISPONIBLE:
        st.caption("Firme con el dedo o el mouse. Se guarda al presionar Guardar.")
        canvas_result = st_canvas(stroke_width=2, stroke_color="#E6EDF3",
                                   background_color="#0D1117", height=160, width=520,
                                   drawing_mode="freedraw", key="firma_canvas",
                                   display_toolbar=True)
        firma_detectada = False
        if canvas_result is not None and canvas_result.image_data is not None:
            import numpy as np
            if canvas_result.image_data[:, :, 0].max() > 100:
                firma_detectada = True
                img_pil = Image.fromarray(canvas_result.image_data.astype("uint8"), "RGBA")
                buf = io.BytesIO()
                img_pil.save(buf, format="PNG")
                st.session_state["firma_png"] = buf.getvalue()
        if firma_detectada:
            st.markdown('<div class="firma-ok">Firma capturada correctamente.</div>',
                        unsafe_allow_html=True)
        else:
            st.markdown('<div class="firma-req">Dibuje su firma en el recuadro.</div>',
                        unsafe_allow_html=True)
    else:
        st.warning("Instale: pip install streamlit-drawable-canvas Pillow")
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Autorización de despacho ───────────────────
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title" style="color:#3FB950">Autorización de Despacho del Vehículo</div>',
                unsafe_allow_html=True)
    desp = st.radio("¿Se autoriza el despacho del vehículo?", options=["Si", "No"],
                     index=0 if st.session_state.get("despacho", "Si") == "Si" else 1,
                     key="despacho_radio", horizontal=True)
    st.session_state["despacho"] = desp
    if desp == "Si":
        st.markdown('<div class="desp-si"><strong style="color:#3FB950">✅ Vehículo autorizado para salir a operación.</strong></div>',
                    unsafe_allow_html=True)
    else:
        st.markdown("""<div class="desp-no">
          <strong style="color:#F85149">🚫 Vehículo RETENIDO.</strong>
          <span style="color:#8B949E;font-size:.85rem"> Al guardar se generará automáticamente una OT correctiva
          (OT-CI-) disponible en el módulo Cierre de OT.</span>
        </div>""", unsafe_allow_html=True)
        if crits == 0:
            st.caption("Nota: no se detectaron ítems críticos en estado Malo. Retención por decisión del prealistador.")
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Progreso ───────────────────────────────────
    total = len(ITEMS_INSPECCION)
    pct = oks / total
    st.markdown(f"""<div class="prog-wrap">
      <div class="prog-lbl">Progreso: {oks}/{total} ítems en estado OK</div>
      <div class="prog-bar"><div class="prog-fill" style="width:{pct*100:.0f}%"></div></div>
    </div>""", unsafe_allow_html=True)

    # ── Validación ─────────────────────────────────
    firma_ok = bool(st.session_state.get("firma_png")) or not CANVAS_DISPONIBLE
    items_malos_sin_modo = [
        item for item, critico, _ in ITEMS_INSPECCION
        if _es_malo(st.session_state["inspeccion"][item]["estado"])
        and not st.session_state["inspeccion"][item]["modo_falla"].strip()
        and MODOS_FALLA_INSP.get(item)
    ]

    avisos = []
    if not prealistador.strip():
        avisos.append("**Nombre del Prealistador**")
    if not numero_interno.strip():
        avisos.append("**Número Interno**")
    if km_int == 0:
        avisos.append("**Kilometraje**")
    if not firma_ok:
        avisos.append("**Firma Digital**")
    if items_malos_sin_modo:
        avisos.append(f"**Modo de falla** en {len(items_malos_sin_modo)} ítem(s) Malo/Mala")
    if avisos:
        st.warning(f"⚠️ Campos obligatorios pendientes: {', '.join(avisos)}")

    if items_malos_sin_modo:
        st.markdown(f"""<div class="error-box">
          ⚠️ <strong>No se puede guardar:</strong> los siguientes ítems están en estado Malo/Mala pero no tienen
          modo de falla seleccionado. Esta información alimenta la OT correctiva:
          <ul style="margin:.4rem 0 0 1rem;color:#F85149">{"".join(f"<li>{item}</li>" for item in items_malos_sin_modo[:5])}
          {"<li style='color:#8B949E'>... y {} más</li>".format(len(items_malos_sin_modo)-5) if len(items_malos_sin_modo) > 5 else ""}
          </ul>
        </div>""", unsafe_allow_html=True)

    campos_ok = (bool(prealistador.strip()) and bool(numero_interno.strip())
                 and km_int > 0
                 and firma_ok and not items_malos_sin_modo)

    cb1, cb2 = st.columns([3, 1])
    with cb1:
        guardar = st.button("💾 GUARDAR INSPECCIÓN", disabled=not campos_ok,
                             use_container_width=True, key="btn_guardar")
    with cb2:
        if st.button("🔄 Limpiar", use_container_width=True, key="btn_limpiar"):
            _reset_form()
            st.rerun()

    # ── GUARDAR ────────────────────────────────────
    if guardar and campos_ok:
        now = datetime.now()
        row = {
            "Fecha_Hora":          now.strftime("%Y-%m-%d %H:%M:%S"),
            "Prealistador":        prealistador.strip(),
            "Numero_Interno":      numero_interno.strip().upper(),
            "Conductor":           conductor.strip(),
            "Kilometraje":         km.strip(),
            "Despacho_Autorizado": st.session_state["despacho"],
        }
        for item_name, es_critico, genero in ITEMS_INSPECCION:
            estado = st.session_state["inspeccion"][item_name]["estado"]
            obs    = st.session_state["inspeccion"][item_name]["obs"].strip()
            modo   = st.session_state["inspeccion"][item_name]["modo_falla"].strip()
            if _es_bueno(estado):
                row[item_name] = estado
            else:
                partes = [estado]
                if modo: partes.append(modo)
                if obs:  partes.append(obs)
                row[item_name] = " | ".join(partes)

        for doc in DOCUMENTOS:
            row[f"Doc_{doc}"] = st.session_state["docs"][doc]

        img_paths = []
        if uploaded_files:
            for f in uploaded_files:
                f.seek(0)
                img_paths.append(_save_image(f, numero_interno.strip()))
        row["Imagenes"] = " | ".join(img_paths) if img_paths else ""

        firma_path = ""
        png_data = st.session_state.get("firma_png")
        if png_data:
            try:
                firma_path = _save_firma_bytes(png_data, numero_interno.strip())
            except Exception as e:
                st.warning(f"No se pudo guardar la firma: {e}")
        row["Firma_Path"] = firma_path

        ot_ci = None
        ots_ci_generadas = []  # Lista de todas las OTs generadas (una por sistema con novedad)
        if st.session_state["despacho"] == "No":
            # Agrupar novedades por sistema estándar (coherente con Reporte de Fallas)
            novedades_por_sistema = _agrupar_novedades_por_sistema(st.session_state["inspeccion"])

            if novedades_por_sistema:
                # Generar UNA OT por cada sistema con novedad
                for sistema, items_nov in novedades_por_sistema.items():
                    ot_sistema = _generar_ot_ci(numero_interno.strip(), sistema)

                    # Construir descripción de la falla: usar modo_falla del ítem más crítico
                    items_criticos = [x for x in items_nov if x[3]]  # es_critico=True
                    items_alertas  = [x for x in items_nov if not x[3]]
                    items_ordenados = items_criticos + items_alertas

                    # Modo de falla principal = del ítem más crítico con modo definido
                    modo_principal = ""
                    for item_name, estado, modo, crit in items_ordenados:
                        if modo:
                            modo_principal = modo
                            break
                    if not modo_principal and items_ordenados:
                        modo_principal = f"{items_ordenados[0][0]}: {items_ordenados[0][1]}"

                    # Resumen de todos los ítems afectados de este sistema
                    resumen_items = []
                    for item_name, estado, modo, crit in items_ordenados:
                        pref = "[CRITICO] " if crit and _es_malo(estado) else "[ALERTA] "
                        resumen_items.append(f"{pref}{item_name}")

                    _registrar_ot_ci(ot_sistema, {
                        "Fecha_Inicio_Inactividad": now.strftime("%Y-%m-%d %H:%M:%S"),
                        "Fecha_Registro_F1":        now.strftime("%Y-%m-%d %H:%M:%S"),
                        "OT":                       ot_sistema,
                        "Numero_Interno":           numero_interno.strip().upper(),
                        "Conductor":                conductor.strip(),
                        "Tipo_Mantenimiento":       "C",
                        "Kilometraje":              km.strip(),
                        "Sistema":                  sistema,
                        "Modo_Falla":               modo_principal,
                        "Items_Afectados":          " | ".join(resumen_items),
                        "Estado_OT":                "P",
                        "Origen":                   "Inspeccion_Preoperacional",
                    })
                    ots_ci_generadas.append((ot_sistema, sistema))

                # Para compatibilidad con pantalla de confirmación: guardar lista
                ot_ci = ots_ci_generadas[0][0] if ots_ci_generadas else None
                row["OT_Correctiva_Generada"] = " | ".join(ot for ot, _ in ots_ci_generadas)

        ok_save, msg = _save_excel(row)
        if ok_save:
            # Auto-actualizar flota_vehiculos.json con el km capturado en la inspección
            accion_flota, km_anterior = _actualizar_flota_desde_inspeccion(
                numero_interno.strip(), km_int
            )
            st.session_state["submitted"]      = True
            st.session_state["saved_data"]     = row
            st.session_state["ot_ci_generada"] = ot_ci
            st.session_state["flota_accion"]   = accion_flota
            st.session_state["flota_km_ant"]   = km_anterior
            st.session_state["contador"]       = st.session_state.get("contador", 0) + 1
            st.session_state["insp_errores_modo"] = []
            st.rerun()
        else:
            st.error(msg)


# ═══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    try:
        st.set_page_config(page_title="Inspección Preoperacional",
                           page_icon="🔍", layout="wide",
                           initial_sidebar_state="collapsed")
    except Exception:
        pass
    run()
