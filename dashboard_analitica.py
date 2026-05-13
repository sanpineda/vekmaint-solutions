"""
═══════════════════════════════════════════════════════════════════
  VEKMAINT SOLUTIONS · Módulo de Analítica
═══════════════════════════════════════════════════════════════════

Dashboard analítico con 4 KPIs operativos y un Pareto de fallas:

  1. Disponibilidad de flota (%)
     = MTBF / (MTBF + MTTR) · MTBF y MTTR calculados solo sobre
       OTs correctivas (CI + CO) cerradas en el mes vigente.

  2. Costo por Kilómetro (CPK ponderado)
     = Σ Costos del mes / Σ Km recorridos del mes
       Mostrado en dos vistas: mes vigente y acumulado año.

  3. % Vehículos no despachados durante inspección preoperacional
     = Inspecciones con Despacho_Autorizado=NO / Total inspecciones del mes × 100

  4. Pareto Top 10 sistemas con más fallas correctivas
     = Cantidad de OT-CI + OT-CO cerradas en el mes vigente, por sistema.

Adicionalmente:
  · Botón "Descargar Hoja de Vida" → Excel con 6 tablas listas para Power BI
  · Botón "Abrir en Power BI" → URL configurable cuando el cliente publique el .pbix
"""
from __future__ import annotations
import os
import io
import json
from datetime import datetime, date, timedelta
from pathlib import Path
import calendar

import streamlit as st
import pandas as pd

# ═══════════════════════════════════════════════════════════════════
#  CONFIG
# ═══════════════════════════════════════════════════════════════════
HISTORICO_MTTO    = "mantenimiento_flotas.xlsx"
HISTORICO_INSP    = "inspecciones_vehiculares.xlsx"
PENDIENTES_DB     = "ots_pendientes.json"
FLOTA_DB          = "flota_vehiculos.json"
CATALOGO_DB       = "catalogo_rutinas.json"
CONFIG_DB         = "config_analitica.json"   # Para guardar la URL del Power BI

HORAS_OPERATIVAS_DIA = 14   # Asunción operativa para flota urbana (parametrizable)

MESES_ES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


# ═══════════════════════════════════════════════════════════════════
#  HELPERS DE PERSISTENCIA
# ═══════════════════════════════════════════════════════════════════
def _cargar_config() -> dict:
    if not os.path.exists(CONFIG_DB):
        return {"power_bi_url": ""}
    try:
        with open(CONFIG_DB, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"power_bi_url": ""}


def _guardar_config(cfg: dict):
    with open(CONFIG_DB, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def _cargar_historico_mtto() -> pd.DataFrame:
    if not os.path.exists(HISTORICO_MTTO):
        return pd.DataFrame()
    try:
        return pd.read_excel(HISTORICO_MTTO)
    except Exception:
        return pd.DataFrame()


def _cargar_historico_insp() -> pd.DataFrame:
    if not os.path.exists(HISTORICO_INSP):
        return pd.DataFrame()
    try:
        return pd.read_excel(HISTORICO_INSP)
    except Exception:
        return pd.DataFrame()


def _cargar_flota() -> dict:
    if not os.path.exists(FLOTA_DB):
        return {}
    try:
        with open(FLOTA_DB, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _cargar_catalogo() -> list:
    if not os.path.exists(CATALOGO_DB):
        return []
    try:
        with open(CATALOGO_DB, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


# ═══════════════════════════════════════════════════════════════════
#  HELPERS DE FECHAS
# ═══════════════════════════════════════════════════════════════════
def _parse_fecha(s):
    """Parsea fecha o datetime en formato string a date."""
    if pd.isna(s) or s == "" or s is None:
        return None
    if isinstance(s, (datetime, pd.Timestamp)):
        return pd.Timestamp(s).to_pydatetime().date()
    if isinstance(s, date):
        return s
    try:
        return datetime.strptime(str(s).split(" ")[0], "%Y-%m-%d").date()
    except Exception:
        return None


def _parse_datetime(s):
    """Parsea fecha+hora en string a datetime."""
    if pd.isna(s) or s == "" or s is None:
        return None
    if isinstance(s, (datetime, pd.Timestamp)):
        return pd.Timestamp(s).to_pydatetime()
    try:
        return datetime.strptime(str(s), "%Y-%m-%d %H:%M:%S")
    except Exception:
        try:
            return datetime.strptime(str(s).split(" ")[0], "%Y-%m-%d")
        except Exception:
            return None


def _rango_mes_vigente():
    hoy = date.today()
    inicio = date(hoy.year, hoy.month, 1)
    ult = calendar.monthrange(hoy.year, hoy.month)[1]
    fin = date(hoy.year, hoy.month, ult)
    return inicio, fin


def _rango_anio_vigente():
    hoy = date.today()
    return date(hoy.year, 1, 1), hoy


# ═══════════════════════════════════════════════════════════════════
#  KPI 1 — DISPONIBILIDAD DE FLOTA (MTBF / MTBF+MTTR)
# ═══════════════════════════════════════════════════════════════════
def calcular_disponibilidad_flota(df_mtto: pd.DataFrame, mes_inicio: date, mes_fin: date):
    """
    Disponibilidad (%) = MTBF / (MTBF + MTTR) × 100

    MTBF (h) y MTTR (h) se calculan SOLO sobre OTs correctivas (CI+CO) cerradas
    en el periodo [mes_inicio, mes_fin].

    MTBF: tiempo operativo medio entre fallas. Asume HORAS_OPERATIVAS_DIA por día.
    MTTR: tiempo medio para reparar (Fecha_Cierre - Fecha_Inicio_Inactividad), en horas.

    Retorna dict con: disponibilidad, mtbf_h, mttr_h, n_correctivos,
                      tiempo_logistica_h, tiempo_reparacion_efectiva_h.
    """
    if df_mtto.empty or "OT" not in df_mtto.columns:
        return None

    # Filtrar solo correctivos cerrados en el mes
    df = df_mtto.copy()
    df["__ot"] = df["OT"].astype(str)
    es_correctiva = df["__ot"].str.startswith("OT-CI-") | df["__ot"].str.startswith("OT-CO-")
    df = df[es_correctiva]

    # Parsear fechas
    df["__fecha_cierre"]  = df.get("Fecha_Cierre_F2", pd.Series([None]*len(df))).apply(_parse_datetime)
    df["__fecha_inicio"]  = df.get("Fecha_Inicio_Inactividad", pd.Series([None]*len(df))).apply(_parse_datetime)

    # Filtro por cierre dentro del mes
    df = df[df["__fecha_cierre"].apply(lambda x: x is not None and mes_inicio <= x.date() <= mes_fin)]

    if df.empty:
        return None

    # Agrupar por OT (en el archivo cada repuesto es una fila, pero tiempos son los mismos)
    df_otu = df.drop_duplicates(subset=["__ot"]).copy()
    n_corr = len(df_otu)

    # MTTR: media de (cierre - inicio inactividad) en horas
    def _delta_h(row):
        ci, fi = row["__fecha_cierre"], row["__fecha_inicio"]
        if ci is None or fi is None:
            return None
        return (ci - fi).total_seconds() / 3600.0
    df_otu["__mttr_h"] = df_otu.apply(_delta_h, axis=1)

    # Tiempo de logística (espera por repuesto) por OT
    def _t_log_h(row):
        d  = row.get("Tiempo_Repuesto_Dias",  0) or 0
        h  = row.get("Tiempo_Repuesto_Horas", 0) or 0
        try:
            return float(d) * 24 + float(h)
        except Exception:
            return 0.0
    df_otu["__tlog_h"] = df_otu.apply(_t_log_h, axis=1)

    # Reparación efectiva por OT = MTTR - tiempo logística (acotado a >= 0)
    df_otu["__treff_h"] = (df_otu["__mttr_h"].fillna(0) - df_otu["__tlog_h"]).clip(lower=0)

    # Promedios sobre las OTs válidas
    mttr_validos = df_otu["__mttr_h"].dropna()
    mttr_h  = float(mttr_validos.mean()) if len(mttr_validos) > 0 else 0.0
    tlog_h  = float(df_otu["__tlog_h"].mean()) if n_corr > 0 else 0.0
    treff_h = float(df_otu["__treff_h"].mean()) if n_corr > 0 else 0.0

    # MTBF: tiempo operativo total / N° fallas
    flota = _cargar_flota()
    n_veh_flota = max(1, len(flota))
    dias_periodo = (mes_fin - mes_inicio).days + 1
    horas_op_totales = n_veh_flota * dias_periodo * HORAS_OPERATIVAS_DIA
    mtbf_h = horas_op_totales / n_corr if n_corr > 0 else 0.0

    # Disponibilidad
    if mtbf_h + mttr_h == 0:
        disp = 0.0
    else:
        disp = (mtbf_h / (mtbf_h + mttr_h)) * 100.0

    return {
        "disponibilidad":   round(disp, 2),
        "mtbf_h":           round(mtbf_h, 1),
        "mttr_h":           round(mttr_h, 1),
        "n_correctivos":    int(n_corr),
        "tiempo_logistica_h":          round(tlog_h, 1),
        "tiempo_reparacion_efectiva_h": round(treff_h, 1),
        "n_vehiculos_flota": n_veh_flota,
        "horas_operativas_dia": HORAS_OPERATIVAS_DIA,
    }


# ═══════════════════════════════════════════════════════════════════
#  KPI 2 — COSTO POR KILÓMETRO (CPK) PONDERADO
# ═══════════════════════════════════════════════════════════════════
def _costos_periodo(df_mtto: pd.DataFrame, ini: date, fin: date) -> float:
    if df_mtto.empty:
        return 0.0
    df = df_mtto.copy()
    df["__cierre"] = df.get("Fecha_Cierre_F2", pd.Series([None]*len(df))).apply(_parse_datetime)
    df = df[df["__cierre"].apply(lambda x: x is not None and ini <= x.date() <= fin)]
    if df.empty:
        return 0.0
    # Sumar costos por OT única (no por fila de repuesto)
    df_unico = df.drop_duplicates(subset=["OT"]) if "OT" in df.columns else df
    costos = pd.to_numeric(df_unico.get("Costo_Total_OT", 0), errors="coerce").fillna(0)
    return float(costos.sum())


def _km_recorridos_periodo(df_insp: pd.DataFrame, ini: date, fin: date):
    """
    Calcula Σ km recorridos por la flota en el periodo, usando el archivo de
    inspecciones preoperacionales:
        km del último día del periodo - km del primer día del periodo
    para cada vehículo, y suma todo.

    Retorna (km_total, detalle_por_vehiculo).
    """
    if df_insp.empty or "Numero_Interno" not in df_insp.columns:
        return 0.0, {}

    df = df_insp.copy()
    df["__fecha"] = df.get("Fecha_Hora", pd.Series([None]*len(df))).apply(_parse_datetime)
    df["__km"]   = pd.to_numeric(df.get("Kilometraje", 0), errors="coerce")
    df = df[df["__fecha"].notna() & df["__km"].notna()]
    df = df[df["__fecha"].apply(lambda x: ini <= x.date() <= fin)]

    if df.empty:
        return 0.0, {}

    detalle = {}
    total = 0.0
    for veh, sub in df.groupby("Numero_Interno"):
        sub_ord = sub.sort_values("__fecha")
        km_ini = float(sub_ord.iloc[0]["__km"])
        km_fin = float(sub_ord.iloc[-1]["__km"])
        km_rec = max(0.0, km_fin - km_ini)
        detalle[veh] = {
            "km_inicial": km_ini,
            "km_final":   km_fin,
            "km_recorridos": km_rec,
            "n_inspecciones": len(sub_ord),
        }
        total += km_rec
    return total, detalle


def calcular_cpk(df_mtto: pd.DataFrame, df_insp: pd.DataFrame):
    """
    Calcula el CPK ponderado de la flota:
        CPK = Σ Costos del periodo / Σ Km recorridos del periodo

    Calcula dos cifras: mes vigente y acumulado año.
    SIEMPRE retorna un dict con los componentes (cost_mes, km_mes, etc.) aunque
    cpk_mes o cpk_ano sean None — para que la UI pueda diagnosticar qué falta.
    """
    mes_ini, mes_fin = _rango_mes_vigente()
    ano_ini, ano_fin = _rango_anio_vigente()

    # Mes vigente
    cost_mes = _costos_periodo(df_mtto, mes_ini, mes_fin)
    km_mes, _ = _km_recorridos_periodo(df_insp, mes_ini, mes_fin)
    cpk_mes = (cost_mes / km_mes) if km_mes > 0 and cost_mes > 0 else None

    # Mes anterior (para tendencia)
    if mes_ini.month == 1:
        ant_ini = date(mes_ini.year - 1, 12, 1)
    else:
        ant_ini = date(mes_ini.year, mes_ini.month - 1, 1)
    ult_ant = calendar.monthrange(ant_ini.year, ant_ini.month)[1]
    ant_fin = date(ant_ini.year, ant_ini.month, ult_ant)
    cost_ant = _costos_periodo(df_mtto, ant_ini, ant_fin)
    km_ant, _ = _km_recorridos_periodo(df_insp, ant_ini, ant_fin)
    cpk_ant = (cost_ant / km_ant) if km_ant > 0 and cost_ant > 0 else None

    # Acumulado año
    cost_ano = _costos_periodo(df_mtto, ano_ini, ano_fin)
    km_ano, _ = _km_recorridos_periodo(df_insp, ano_ini, ano_fin)
    cpk_ano = (cost_ano / km_ano) if km_ano > 0 and cost_ano > 0 else None

    # Tendencia mes vs mes anterior
    tendencia = None
    if cpk_mes is not None and cpk_ant is not None and cpk_ant > 0:
        tendencia = ((cpk_mes - cpk_ant) / cpk_ant) * 100

    return {
        "cpk_mes":      cpk_mes,
        "cost_mes":     cost_mes,
        "km_mes":       km_mes,
        "cpk_ano":      cpk_ano,
        "cost_ano":     cost_ano,
        "km_ano":       km_ano,
        "cpk_anterior": cpk_ant,
        "tendencia_pct": tendencia,
        "nombre_mes":   MESES_ES[mes_ini.month],
        "ano":          mes_ini.year,
    }


# ═══════════════════════════════════════════════════════════════════
#  KPI 3 — % VEHÍCULOS NO DESPACHADOS
# ═══════════════════════════════════════════════════════════════════
def calcular_no_despachados(df_insp: pd.DataFrame, df_mtto: pd.DataFrame,
                            mes_inicio: date, mes_fin: date):
    """
    % No Despachados = Inspecciones con Despacho_Autorizado=NO / Total inspecciones del mes × 100
    """
    if df_insp.empty:
        return None

    df = df_insp.copy()
    df["__fecha"] = df.get("Fecha_Hora", pd.Series([None]*len(df))).apply(_parse_datetime)
    df = df[df["__fecha"].apply(lambda x: x is not None and mes_inicio <= x.date() <= mes_fin)]

    n_total = len(df)
    if n_total == 0:
        return None

    despacho_col = df.get("Despacho_Autorizado", pd.Series([""]*n_total)).astype(str).str.strip().str.upper()
    no_despachadas = despacho_col.str.startswith("NO")
    n_no = int(no_despachadas.sum())

    pct = (n_no / n_total) * 100 if n_total > 0 else 0
    return {
        "pct":               round(pct, 1),
        "n_no_despachadas":  n_no,
        "n_total_inspecc":   n_total,
        "nombre_mes":        MESES_ES[mes_inicio.month],
    }


# ═══════════════════════════════════════════════════════════════════
#  KPI 4 — DISTRIBUCIÓN PREVENTIVO vs CORRECTIVO (POR COSTO)
# ═══════════════════════════════════════════════════════════════════
def calcular_distribucion_pyc(df_mtto: pd.DataFrame, mes_inicio: date, mes_fin: date):
    """
    Calcula la distribución del costo de mantenimiento entre preventivo y correctivo.

    Categorización:
      · Preventivo: OT-P (preventivo programado) + OT-M (mayor planificado)
      · Correctivo: OT-CI (correctivo de inspección) + OT-CO (correctivo de operación)

    El indicador mide la madurez de la gestión de mantenimiento:
      · ALTA   ≥ 60% costo preventivo (gestión proactiva)
      · MEDIA  40-60% (en transición)
      · BAJA   < 40% (modo "apaga incendios")

    Calcula también la tendencia comparando contra el mes anterior.

    Retorna dict con:
      pct_preventivo, pct_correctivo, costo_preventivo, costo_correctivo,
      n_preventivos, n_correctivos, n_p, n_m, n_ci, n_co,
      madurez (str), tendencia_pct (None o float vs mes anterior).
    """
    if df_mtto.empty or "OT" not in df_mtto.columns:
        return None

    def _calcular_periodo(ini: date, fin: date):
        """Calcula costos y conteos del periodo. Retorna dict o None si no hay datos."""
        df = df_mtto.copy()
        df["__ot"] = df["OT"].astype(str)
        df["__cierre"] = df.get("Fecha_Cierre_F2", pd.Series([None]*len(df))).apply(_parse_datetime)
        df = df[df["__cierre"].apply(lambda x: x is not None and ini <= x.date() <= fin)]
        if df.empty:
            return None
        # Una OT por fila para conteos y costos (no por repuesto)
        df_unico = df.drop_duplicates(subset=["__ot"]).copy()

        # Categorizar
        es_p  = df_unico["__ot"].str.startswith("OT-P-")
        es_m  = df_unico["__ot"].str.startswith("OT-M-")
        es_ci = df_unico["__ot"].str.startswith("OT-CI-")
        es_co = df_unico["__ot"].str.startswith("OT-CO-")

        df_unico["__costo"] = pd.to_numeric(df_unico.get("Costo_Total_OT", 0), errors="coerce").fillna(0)

        cost_prev = float(df_unico.loc[es_p | es_m, "__costo"].sum())
        cost_corr = float(df_unico.loc[es_ci | es_co, "__costo"].sum())
        n_p  = int(es_p.sum())
        n_m  = int(es_m.sum())
        n_ci = int(es_ci.sum())
        n_co = int(es_co.sum())

        return {
            "cost_prev": cost_prev, "cost_corr": cost_corr,
            "n_p": n_p, "n_m": n_m, "n_ci": n_ci, "n_co": n_co,
            "n_prev": n_p + n_m, "n_corr": n_ci + n_co,
        }

    actual = _calcular_periodo(mes_inicio, mes_fin)
    if actual is None:
        return None

    cost_total = actual["cost_prev"] + actual["cost_corr"]
    if cost_total == 0:
        # Hay OTs cerradas pero todas con costo $0 — no se puede calcular distribución
        return None

    pct_p = (actual["cost_prev"] / cost_total) * 100
    pct_c = (actual["cost_corr"] / cost_total) * 100

    # Madurez según el ratio preventivo
    if pct_p >= 60:   madurez = "ALTA"
    elif pct_p >= 40: madurez = "MEDIA"
    else:              madurez = "BAJA"

    # Tendencia vs mes anterior
    if mes_inicio.month == 1:
        ant_ini = date(mes_inicio.year - 1, 12, 1)
    else:
        ant_ini = date(mes_inicio.year, mes_inicio.month - 1, 1)
    ult_ant = calendar.monthrange(ant_ini.year, ant_ini.month)[1]
    ant_fin = date(ant_ini.year, ant_ini.month, ult_ant)
    anterior = _calcular_periodo(ant_ini, ant_fin)
    tendencia = None
    if anterior is not None:
        cost_total_ant = anterior["cost_prev"] + anterior["cost_corr"]
        if cost_total_ant > 0:
            pct_p_ant = (anterior["cost_prev"] / cost_total_ant) * 100
            tendencia = pct_p - pct_p_ant   # diferencia en puntos porcentuales

    return {
        "pct_preventivo":   round(pct_p, 1),
        "pct_correctivo":   round(pct_c, 1),
        "costo_preventivo": actual["cost_prev"],
        "costo_correctivo": actual["cost_corr"],
        "costo_total":      cost_total,
        "n_preventivos":    actual["n_prev"],
        "n_correctivos":    actual["n_corr"],
        "n_p":              actual["n_p"],
        "n_m":              actual["n_m"],
        "n_ci":             actual["n_ci"],
        "n_co":             actual["n_co"],
        "madurez":          madurez,
        "tendencia_pp":     tendencia,   # diferencia en puntos porcentuales
    }


# ═══════════════════════════════════════════════════════════════════
#  KPI 5 — PARETO DE FALLAS (Top 10 sistemas por cantidad de OT)
# ═══════════════════════════════════════════════════════════════════
# (la marca anterior estaba duplicada — eliminada)


def calcular_pareto_fallas(df_mtto: pd.DataFrame, mes_inicio: date, mes_fin: date) -> list:
    """
    Top 10 sistemas con más OTs correctivas (CI+CO) cerradas en el mes vigente.
    Retorna lista de dicts ordenados desc por cantidad: [{sistema, n_ots, pct, pct_acum}].
    """
    if df_mtto.empty or "OT" not in df_mtto.columns:
        return []

    df = df_mtto.copy()
    df["__ot"] = df["OT"].astype(str)
    es_correctiva = df["__ot"].str.startswith("OT-CI-") | df["__ot"].str.startswith("OT-CO-")
    df = df[es_correctiva]
    df["__cierre"] = df.get("Fecha_Cierre_F2", pd.Series([None]*len(df))).apply(_parse_datetime)
    df = df[df["__cierre"].apply(lambda x: x is not None and mes_inicio <= x.date() <= mes_fin)]

    if df.empty:
        return []

    # Una OT por fila (no por repuesto)
    df_unico = df.drop_duplicates(subset=["__ot"])

    conteo = df_unico.get("Sistema", pd.Series([])).fillna("Sin clasificar").value_counts()
    total = int(conteo.sum())
    if total == 0:
        return []

    pareto = []
    pct_acum = 0.0
    for sistema, n in conteo.head(10).items():
        pct = (n / total) * 100
        pct_acum += pct
        pareto.append({
            "sistema":  str(sistema),
            "n_ots":    int(n),
            "pct":      round(pct, 1),
            "pct_acum": round(pct_acum, 1),
        })
    return pareto


# ═══════════════════════════════════════════════════════════════════
#  HOJA DE VIDA — EXPORTACIÓN A EXCEL CON 5 TABLAS
# ═══════════════════════════════════════════════════════════════════
def generar_hoja_de_vida() -> bytes:
    """
    Genera un archivo Excel con 6 hojas listas para integrar con Power BI:
      1. Maestro_Vehiculos          — datos básicos de la flota
      2. Catalogo_Rutinas           — rutinas preventivas con costo estándar
      3. Inspecciones               — formato ANCHO (igual que el archivo histórico),
                                       con coloreado condicional: rojo para "Malo",
                                       amarillo para "Regular", celdas blancas para "Bueno".
                                       Hipervínculos a firmas y fotos. Apto para revisión humana.
      4. Inspecciones_Items         — formato LARGO (una fila por inspección × ítem),
                                       con columnas: Inspeccion_ID, Fecha, Vehículo, Item,
                                       Estado (normalizado), Modo_Falla, Tiene_Novedad.
                                       Apto para Power BI sin transformaciones adicionales.
      5. Mantenimientos             — todas las OTs cerradas con tipo categorizado
      6. Cumplimiento_Preventivo    — derivada con días delta y flag a tiempo
    """
    output = io.BytesIO()

    # ── Tabla 1: Maestro_Vehiculos ──
    flota = _cargar_flota()
    rows_maestro = []
    for veh_id, vdata in flota.items():
        rows_maestro.append({
            "Numero_Interno": veh_id,
            "Marca":          vdata.get("marca", ""),
            "Modelo":         vdata.get("modelo", ""),
            "Tipo_Vehiculo":  vdata.get("tipo", ""),
            "Km_Actual":      vdata.get("km_actual", 0),
            "Fecha_Ultima_Inspeccion": vdata.get("fecha_ultima_inspeccion", ""),
        })
    df_maestro = pd.DataFrame(rows_maestro) if rows_maestro else pd.DataFrame(
        columns=["Numero_Interno","Marca","Modelo","Tipo_Vehiculo","Km_Actual","Fecha_Ultima_Inspeccion"])

    # ── Tabla 2: Catalogo_Rutinas ──
    catalogo = _cargar_catalogo()
    rows_cat = []
    for r in catalogo:
        costo_repuestos = sum(rep.get("cantidad", 0) * rep.get("costo_unit", 0)
                                for rep in r.get("repuestos", []))
        costo_estandar = r.get("mano_obra", 0) + costo_repuestos
        rows_cat.append({
            "Rutina_ID":         r.get("id", ""),
            "Nombre":            r.get("nombre", ""),
            "Sistema":           r.get("sistema", ""),
            "Periodicidad_Km":   r.get("periodicidad_km", 0),
            "Periodicidad_Dias": r.get("periodicidad_dias", 0),
            "Duracion_Horas":    r.get("duracion_horas", 0),
            "Costo_Mano_Obra":   r.get("mano_obra", 0),
            "Costo_Repuestos_Estandar": costo_repuestos,
            "Costo_Estandar_Total":     costo_estandar,
        })
    df_cat = pd.DataFrame(rows_cat) if rows_cat else pd.DataFrame(
        columns=["Rutina_ID","Nombre","Sistema","Periodicidad_Km","Periodicidad_Dias",
                 "Duracion_Horas","Costo_Mano_Obra","Costo_Repuestos_Estandar","Costo_Estandar_Total"])

    # ── Tabla 3a: Inspecciones (formato ANCHO original — humano) ──
    # ── Tabla 3b: Inspecciones_Items (formato LARGO — para Power BI) ──
    df_insp_raw = _cargar_historico_insp()
    if not df_insp_raw.empty:
        # Identificar columnas básicas (metadata) e ítems (estados de cada componente)
        cols_basicas = [c for c in ["Fecha_Hora", "Prealistador", "Numero_Interno",
                                       "Conductor", "Kilometraje", "Despacho_Autorizado"]
                          if c in df_insp_raw.columns]
        cols_doc = [c for c in df_insp_raw.columns if c.startswith("Doc_") or c == "Imagenes"]
        items_cols = [c for c in df_insp_raw.columns
                       if c not in cols_basicas and c not in cols_doc]

        # 3a — Hoja ancha: igual que el archivo original (todos los ítems como columnas).
        # Mantenemos el orden exacto: básicas + ítems + documentos.
        col_order = cols_basicas + items_cols + cols_doc
        df_insp_wide = df_insp_raw[[c for c in col_order if c in df_insp_raw.columns]].copy()

        # 3b — Hoja larga: una fila por (inspección × ítem).
        # Solo se incluyen los ítems que tienen valor no vacío.
        rows_long = []
        for idx, row in df_insp_raw.iterrows():
            inspeccion_id = f"INS-{idx+1:05d}"
            fecha = row.get("Fecha_Hora", "")
            veh   = row.get("Numero_Interno", "")
            preal = row.get("Prealistador", "")
            for c in items_cols:
                valor = str(row.get(c, "") or "").strip()
                if not valor:
                    continue
                # Separar estado y modo de falla si vienen concatenados
                # Formato típico almacenado: "Bueno" / "Regular | <modo>" / "Malo | <modo>" / "Regular - <modo>"
                if " | " in valor:
                    estado, modo = valor.split(" | ", 1)
                elif " - " in valor and not valor.lower().startswith("bueno"):
                    estado, modo = valor.split(" - ", 1)
                else:
                    estado, modo = valor, ""
                estado = estado.strip()
                # Normalizar estado para análisis
                est_low = estado.lower()
                if est_low.startswith(("bueno", "buena")):  est_norm = "Bueno"
                elif est_low.startswith("regular"):           est_norm = "Regular"
                elif est_low.startswith(("malo", "mala")):    est_norm = "Malo"
                else:                                          est_norm = estado
                rows_long.append({
                    "Inspeccion_ID":    inspeccion_id,
                    "Fecha_Hora":       fecha,
                    "Numero_Interno":   veh,
                    "Prealistador":     preal,
                    "Item":             c,
                    "Estado":           est_norm,
                    "Modo_Falla":       modo.strip(),
                    "Tiene_Novedad":    "Sí" if est_norm != "Bueno" else "No",
                })
        df_insp_long = pd.DataFrame(rows_long) if rows_long else pd.DataFrame(
            columns=["Inspeccion_ID","Fecha_Hora","Numero_Interno","Prealistador",
                     "Item","Estado","Modo_Falla","Tiene_Novedad"])
    else:
        df_insp_wide = pd.DataFrame(columns=["Fecha_Hora","Prealistador","Numero_Interno",
                                              "Conductor","Kilometraje","Despacho_Autorizado"])
        df_insp_long = pd.DataFrame(columns=["Inspeccion_ID","Fecha_Hora","Numero_Interno",
                                              "Prealistador","Item","Estado","Modo_Falla",
                                              "Tiene_Novedad"])
        cols_basicas = []
        items_cols = []
        cols_doc = []

    # ── Tabla 4: Mantenimientos ──
    df_mtto_raw = _cargar_historico_mtto()
    if not df_mtto_raw.empty:
        df_mtto = df_mtto_raw.copy()
        def _tipo_ot(ot):
            ot = str(ot)
            if ot.startswith("OT-CI-"): return "Correctivo Inspección"
            if ot.startswith("OT-CO-"): return "Correctivo Operación"
            if ot.startswith("OT-P-"):  return "Preventivo"
            if ot.startswith("OT-M-"):  return "Mayor"
            return "Otro"
        df_mtto["Tipo_OT_Categoria"] = df_mtto.get("OT", "").apply(_tipo_ot)
    else:
        df_mtto = pd.DataFrame()

    # ── Tabla 5: Cumplimiento_Preventivo ──
    rows_cumpl = []
    if not df_mtto_raw.empty and "OT" in df_mtto_raw.columns:
        df_p = df_mtto_raw[df_mtto_raw["OT"].astype(str).str.startswith("OT-P-")]
        df_p_unico = df_p.drop_duplicates(subset=["OT"])
        for _, row in df_p_unico.iterrows():
            ot = str(row.get("OT", ""))
            f_prog_str = row.get("Fecha_Programada", "")
            f_cierre_str = row.get("Fecha_Cierre_F2", "")
            f_prog = _parse_fecha(f_prog_str)
            f_cierre = _parse_fecha(f_cierre_str)
            if f_prog is None or f_cierre is None:
                continue
            delta = (f_cierre - f_prog).days
            cumple = (-5 <= delta <= 2)
            rows_cumpl.append({
                "OT":               ot,
                "Numero_Interno":   row.get("Numero_Interno", ""),
                "Sistema":          row.get("Sistema", ""),
                "Fecha_Programada": f_prog.strftime("%Y-%m-%d"),
                "Fecha_Cierre":     f_cierre.strftime("%Y-%m-%d"),
                "Dias_Delta":       delta,
                "Cumplio_A_Tiempo": "Sí" if cumple else "No",
                "Costo_Total":      row.get("Costo_Total_OT", 0),
            })
    df_cumpl = pd.DataFrame(rows_cumpl) if rows_cumpl else pd.DataFrame(
        columns=["OT", "Numero_Interno", "Sistema", "Fecha_Programada", "Fecha_Cierre",
                  "Dias_Delta", "Cumplio_A_Tiempo", "Costo_Total"]
    )

    # Escribir todas las hojas
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_maestro    .to_excel(writer, sheet_name="Maestro_Vehiculos",       index=False)
        df_cat        .to_excel(writer, sheet_name="Catalogo_Rutinas",        index=False)
        df_insp_wide  .to_excel(writer, sheet_name="Inspecciones",            index=False)
        df_insp_long  .to_excel(writer, sheet_name="Inspecciones_Items",      index=False)
        df_mtto       .to_excel(writer, sheet_name="Mantenimientos",          index=False)
        df_cumpl      .to_excel(writer, sheet_name="Cumplimiento_Preventivo", index=False)

        # ── Formato condicional + estilos en hojas ──
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        fill_header  = PatternFill("solid", fgColor="0A1628")
        fill_red     = PatternFill("solid", fgColor="FFCCCC")
        fill_yellow  = PatternFill("solid", fgColor="FFF3CC")
        fill_green   = PatternFill("solid", fgColor="DFF5DD")
        font_header  = Font(color="FFFFFF", bold=True, size=10)
        font_link    = Font(color="0563C1", underline="single", size=10)

        def _aplicar_header(ws):
            for cell in ws[1]:
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

        def _autoajustar(ws, max_width=45):
            for col in ws.columns:
                lens = [len(str(c.value)) for c in col if c.value is not None]
                ml = max(lens) if lens else 10
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, max_width)

        # ── Hoja Inspecciones (formato ANCHO con coloreado por estado) ──
        ws_insp = writer.sheets["Inspecciones"]
        _aplicar_header(ws_insp)
        # Identificar columnas de ítems (las que no son básicas ni Doc_)
        nombres_basicas = set(cols_basicas)
        nombres_doc = set(cols_doc)
        # Recorremos cada fila y aplicamos color a las celdas de ítems
        for row_idx, row in enumerate(ws_insp.iter_rows(min_row=2, max_row=ws_insp.max_row), start=2):
            for cell in row:
                col_name = ws_insp.cell(row=1, column=cell.column).value
                if col_name in nombres_basicas:
                    continue
                if col_name in nombres_doc:
                    # Convertir documento en hipervínculo si es ruta
                    if cell.value and isinstance(cell.value, str) and cell.value.strip():
                        # Múltiples imágenes separadas por |
                        rutas = [r.strip() for r in cell.value.split("|") if r.strip()]
                        if len(rutas) == 1:
                            try:
                                cell.hyperlink = rutas[0]
                                cell.font = font_link
                            except Exception:
                                pass
                        # Si hay múltiples, dejamos como texto (Excel solo soporta 1 hyperlink/celda)
                    continue
                # Es columna de ítem — aplicar coloreado
                if cell.value and isinstance(cell.value, str):
                    v = cell.value.strip().lower()
                    if v.startswith("malo") or v.startswith("mala"):
                        cell.fill = fill_red
                    elif v.startswith("regular"):
                        cell.fill = fill_yellow
                    elif v.startswith("bueno") or v.startswith("buena"):
                        # Sin pintar — celda blanca para que destaquen las novedades
                        pass
        _autoajustar(ws_insp)
        ws_insp.freeze_panes = "B2"  # Congelar primera fila + primera columna

        # ── Hoja Inspecciones_Items (formato LARGO para Power BI) ──
        ws_long = writer.sheets["Inspecciones_Items"]
        _aplicar_header(ws_long)
        # Colorear la columna Estado para que también sea visualmente útil
        col_estado_idx = None
        col_novedad_idx = None
        for c in ws_long[1]:
            if c.value == "Estado":         col_estado_idx = c.column
            if c.value == "Tiene_Novedad":  col_novedad_idx = c.column
        if col_estado_idx:
            for row in ws_long.iter_rows(min_row=2, max_row=ws_long.max_row):
                est_cell = ws_long.cell(row=row[0].row, column=col_estado_idx)
                v = (est_cell.value or "").lower()
                if v == "malo":      est_cell.fill = fill_red
                elif v == "regular": est_cell.fill = fill_yellow
                elif v == "bueno":   est_cell.fill = fill_green
        _autoajustar(ws_long, max_width=35)
        ws_long.freeze_panes = "A2"

        # ── Header en las demás hojas ──
        for sheet_name in ["Maestro_Vehiculos", "Catalogo_Rutinas",
                            "Mantenimientos", "Cumplimiento_Preventivo"]:
            ws = writer.sheets[sheet_name]
            _aplicar_header(ws)
            _autoajustar(ws)
            ws.freeze_panes = "A2"

    output.seek(0)
    return output.getvalue()


# ═══════════════════════════════════════════════════════════════════
#  RENDER PRINCIPAL
# ═══════════════════════════════════════════════════════════════════
def run():
    # ── Estilos del módulo ──
    st.markdown("""<style>
    :root{
      --bg:#0D1117;--bg-deep:#080C14;--surface:#161B22;--surface2:#1F2937;
      --border:#30363D;--text:#E6EDF3;--muted:#8B949E;
      --teal:#3FE0C6;--teal-dk:#1FA38B;--teal-lt:#A5F0E2;
      --green:#3FB950;--orange:#FF6B00;--red:#F85149;--blue:#0A84FF;--purple:#7C4DFF;
    }
    .ana-header{display:flex;align-items:center;gap:1rem;padding:1.2rem 1.5rem;
      background:linear-gradient(135deg, rgba(63,224,198,.18) 0%, rgba(31,163,139,.08) 100%);
      border:1px solid rgba(63,224,198,.4);border-left:5px solid var(--teal);
      border-radius:14px;margin-bottom:1.2rem}
    .ana-icon{font-size:2.2rem}
    .ana-title{font-family:'Exo 2',sans-serif;font-size:1.6rem;font-weight:700;color:var(--teal-lt);margin:0;letter-spacing:.5px}
    .ana-sub{color:var(--muted);font-size:.85rem;margin-top:.15rem}

    .kpi-card{background:var(--surface);border:1px solid var(--border);border-radius:12px;
      padding:1.1rem 1.3rem;height:100%}
    .kpi-card-head{display:flex;align-items:center;gap:.5rem;color:var(--muted);
      font-size:.78rem;text-transform:uppercase;letter-spacing:.5px;font-weight:600;margin-bottom:.5rem}
    .kpi-value{font-family:'Exo 2',sans-serif;font-size:2.6rem;font-weight:800;line-height:1.1;color:var(--text)}
    .kpi-unit{font-size:1.1rem;color:var(--muted);margin-left:.3rem}
    .kpi-sub{font-size:.78rem;color:var(--muted);margin-top:.5rem;line-height:1.4}
    .kpi-sub strong{color:var(--text)}
    .kpi-trend-up{color:var(--red);font-size:.78rem}
    .kpi-trend-down{color:var(--green);font-size:.78rem}
    .kpi-trend-flat{color:var(--muted);font-size:.78rem}

    .kpi-good   {border-left:5px solid var(--green)}
    .kpi-warn   {border-left:5px solid var(--orange)}
    .kpi-bad    {border-left:5px solid var(--red)}
    .kpi-info   {border-left:5px solid var(--teal)}
    .kpi-neutral{border-left:5px solid var(--purple)}

    .pareto-row{display:grid;grid-template-columns:1fr 80px;gap:.6rem;align-items:center;
      padding:.55rem .7rem;border-bottom:1px solid var(--border)}
    .pareto-row:last-child{border-bottom:none}
    .pareto-bar{position:relative;background:var(--surface2);border-radius:6px;height:22px;overflow:hidden}
    .pareto-bar-fill{height:100%;background:linear-gradient(90deg, var(--orange) 0%, #FF9A40 100%);
      border-radius:6px;display:flex;align-items:center;padding-left:.5rem;color:#000;font-weight:700;font-size:.75rem}
    .pareto-label{font-size:.85rem;color:var(--text);font-weight:500}
    .pareto-count{font-size:.85rem;color:var(--muted);text-align:right}
    .pareto-pct  {font-size:.7rem;color:var(--muted);text-align:right}

    .info-banner{background:rgba(10,132,255,.08);border:1px solid rgba(10,132,255,.3);
      border-radius:10px;padding:.7rem 1rem;font-size:.78rem;color:#7EC4FF;margin:1rem 0}

    .empty-state{text-align:center;padding:2rem 1rem;color:var(--muted)}
    .empty-state .em-icon{font-size:2.5rem;margin-bottom:.5rem;opacity:.5}
    </style>""", unsafe_allow_html=True)

    # Botón volver al hub (usa la misma clave de estado que los demás módulos)
    cb1, _ = st.columns([1, 8])
    with cb1:
        if st.button("← Volver", key="ana_volver", use_container_width=True):
            st.session_state["modulo"] = "hub"
            try:
                st.query_params.clear()
            except Exception:
                pass
            st.rerun()

    # Header
    mes_ini, mes_fin = _rango_mes_vigente()
    st.markdown(f"""<div class="ana-header">
      <div class="ana-icon">📊</div>
      <div>
        <h1 class="ana-title">Dashboard Analítico · KPIs</h1>
        <div class="ana-sub">Indicadores estratégicos · {MESES_ES[mes_ini.month]} {mes_ini.year} · flota completa</div>
      </div>
    </div>""", unsafe_allow_html=True)

    # ── Cargar datos ──
    df_mtto = _cargar_historico_mtto()
    df_insp = _cargar_historico_insp()

    # ═══════════════════════════════════════════════════════
    #  FILA 1: KPIs PRINCIPALES (3 indicadores)
    # ═══════════════════════════════════════════════════════
    col1, col2, col3 = st.columns(3, gap="medium")

    # ── KPI 1: Disponibilidad ──
    with col1:
        disp_data = calcular_disponibilidad_flota(df_mtto, mes_ini, mes_fin)
        if disp_data is None:
            st.markdown("""<div class="kpi-card kpi-neutral">
              <div class="kpi-card-head">🟢 Disponibilidad de Flota</div>
              <div class="kpi-value">—<span class="kpi-unit">%</span></div>
              <div class="kpi-sub">Sin OTs correctivas cerradas en el mes vigente
                — la disponibilidad se calcula con base en MTBF y MTTR de correctivos.</div>
            </div>""", unsafe_allow_html=True)
        else:
            d = disp_data["disponibilidad"]
            cls = "kpi-good" if d >= 95 else ("kpi-warn" if d >= 85 else "kpi-bad")
            st.markdown(f"""<div class="kpi-card {cls}">
              <div class="kpi-card-head">🟢 Disponibilidad de Flota</div>
              <div class="kpi-value">{d:.1f}<span class="kpi-unit">%</span></div>
              <div class="kpi-sub">
                MTBF: <strong>{disp_data['mtbf_h']:.0f} h</strong> ·
                MTTR: <strong>{disp_data['mttr_h']:.1f} h</strong><br>
                <span style="color:var(--muted)">
                  {disp_data['n_correctivos']} OT correctiva{'s' if disp_data['n_correctivos']!=1 else ''} ·
                  Logística: {disp_data['tiempo_logistica_h']:.1f} h ·
                  Reparación efectiva: {disp_data['tiempo_reparacion_efectiva_h']:.1f} h
                </span>
              </div>
            </div>""", unsafe_allow_html=True)

    # ── KPI 2: CPK ──
    with col2:
        cpk_data = calcular_cpk(df_mtto, df_insp)
        # cpk_data SIEMPRE es no-None ahora (puede tener cpk_mes=None pero trae cost_mes y km_mes)
        if cpk_data is None:
            cost_mes = 0
            km_mes = 0
            cpk_mes = None
            cpk_ano = None
            nombre_mes = MESES_ES[mes_ini.month]
            ano_actual = mes_ini.year
            tend = None
        else:
            cost_mes = cpk_data["cost_mes"]
            km_mes = cpk_data["km_mes"]
            cpk_mes = cpk_data["cpk_mes"]
            cpk_ano = cpk_data["cpk_ano"]
            nombre_mes = cpk_data["nombre_mes"]
            ano_actual = cpk_data["ano"]
            tend = cpk_data["tendencia_pct"]

        # Diagnóstico claro de qué falta
        falta_costos = (cost_mes == 0)
        falta_km     = (km_mes == 0)

        if cpk_mes is None:
            # Determinar mensaje específico según qué falta
            if falta_costos and falta_km:
                msg_falta = "Sin OTs cerradas ni inspecciones con km en el mes."
            elif falta_costos:
                msg_falta = "Sin OTs cerradas en el mes (todas pendientes de cierre)."
            elif falta_km:
                msg_falta = "Sin km recorridos registrados (se requieren ≥2 inspecciones por vehículo en días distintos)."
            else:
                msg_falta = "Datos insuficientes para el cálculo."

            cost_str = f"${cost_mes:,.0f}".replace(",", ".") if cost_mes else "$ 0"
            km_str = f"{km_mes:,.0f}".replace(",", ".") if km_mes else "0"
            st.markdown(f"""<div class="kpi-card kpi-neutral">
              <div class="kpi-card-head">💵 Costo por Kilómetro (CPK)</div>
              <div class="kpi-value">—<span class="kpi-unit">/km</span></div>
              <div class="kpi-sub">
                <span style="color:#FF9A40">⚠️ {msg_falta}</span><br>
                <span style="color:var(--muted)">
                  Costos del mes: <strong>{cost_str}</strong> · Km recorridos: <strong>{km_str}</strong>
                </span>
              </div>
            </div>""", unsafe_allow_html=True)
        else:
            cpk_mes_str = f"${cpk_mes:,.0f}".replace(",", ".")
            cpk_ano_str = f"${cpk_ano:,.0f}".replace(",", ".") if cpk_ano is not None else "—"

            if tend is None:
                tend_html = '<span class="kpi-trend-flat">— sin mes anterior</span>'
            elif tend > 1:
                tend_html = f'<span class="kpi-trend-up">▲ {tend:.1f}% vs mes anterior</span>'
            elif tend < -1:
                tend_html = f'<span class="kpi-trend-down">▼ {abs(tend):.1f}% vs mes anterior</span>'
            else:
                tend_html = '<span class="kpi-trend-flat">≈ estable vs mes anterior</span>'

            cost_mes_str = f"${cost_mes:,.0f}".replace(",", ".")
            km_mes_str = f"{km_mes:,.0f}".replace(",", ".")

            st.markdown(f"""<div class="kpi-card kpi-info">
              <div class="kpi-card-head">💵 Costo por Kilómetro (CPK ponderado)</div>
              <div class="kpi-value">{cpk_mes_str}<span class="kpi-unit">/km</span></div>
              <div style="font-size:.78rem;color:var(--muted);margin-top:.1rem">
                {nombre_mes} · acumulado {ano_actual}: <strong style="color:var(--text)">{cpk_ano_str}/km</strong>
              </div>
              <div class="kpi-sub">
                {tend_html}<br>
                <span style="color:var(--muted)">
                  Mes: {cost_mes_str} / {km_mes_str} km
                </span>
              </div>
            </div>""", unsafe_allow_html=True)

    # ── KPI 3: % No Despachados ──
    with col3:
        nodes = calcular_no_despachados(df_insp, df_mtto, mes_ini, mes_fin)
        if nodes is None:
            st.markdown("""<div class="kpi-card kpi-neutral">
              <div class="kpi-card-head">🚫 Vehículos no despachados</div>
              <div class="kpi-value">—<span class="kpi-unit">%</span></div>
              <div class="kpi-sub">Sin inspecciones preoperacionales registradas
                en el mes vigente.</div>
            </div>""", unsafe_allow_html=True)
        else:
            pct = nodes["pct"]
            cls = "kpi-good" if pct < 3 else ("kpi-warn" if pct < 7 else "kpi-bad")
            st.markdown(f"""<div class="kpi-card {cls}">
              <div class="kpi-card-head">🚫 Vehículos no despachados (inspección)</div>
              <div class="kpi-value">{pct:.1f}<span class="kpi-unit">%</span></div>
              <div class="kpi-sub">
                <strong>{nodes['n_no_despachadas']}</strong> de
                <strong>{nodes['n_total_inspecc']}</strong> inspecciones
                bloquearon el despacho<br>
                <span style="color:var(--muted)">
                  Mide la efectividad del filtro preoperacional sobre el riesgo operativo.
                </span>
              </div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════
    #  KPI 4: DISTRIBUCIÓN PREVENTIVO vs CORRECTIVO (full-width)
    # ═══════════════════════════════════════════════════════
    st.markdown("""<h2 style="font-family:'Exo 2',sans-serif;color:var(--teal-lt);
        font-size:1.2rem;margin:0 0 .8rem 0;letter-spacing:.5px">
        ⚖️ Distribución del Costo: Preventivo vs Correctivo
        </h2>""", unsafe_allow_html=True)

    dist = calcular_distribucion_pyc(df_mtto, mes_ini, mes_fin)
    if dist is None:
        st.markdown("""<div class="kpi-card kpi-neutral">
          <div class="empty-state">
            <div class="em-icon">⚖️</div>
            <div>No hay OTs cerradas con costo registrado en el mes vigente.</div>
            <div style="font-size:.78rem;margin-top:.3rem">
              Este indicador requiere al menos una OT cerrada (preventiva o correctiva)
              con su costo total registrado.
            </div>
          </div>
        </div>""", unsafe_allow_html=True)
    else:
        # Color del KPI según madurez
        if   dist["madurez"] == "ALTA":  cls_card, color_madurez = "kpi-good", "var(--green)"
        elif dist["madurez"] == "MEDIA": cls_card, color_madurez = "kpi-warn", "var(--orange)"
        else:                              cls_card, color_madurez = "kpi-bad",  "var(--red)"

        # Tendencia vs mes anterior (en puntos porcentuales)
        tend_pp = dist["tendencia_pp"]
        if tend_pp is None:
            tend_html = '<span class="kpi-trend-flat">— sin mes anterior</span>'
        elif tend_pp > 1:
            tend_html = f'<span class="kpi-trend-down">▲ +{tend_pp:.1f} pp en preventivo vs mes anterior</span>'
        elif tend_pp < -1:
            tend_html = f'<span class="kpi-trend-up">▼ -{abs(tend_pp):.1f} pp en preventivo vs mes anterior</span>'
        else:
            tend_html = '<span class="kpi-trend-flat">≈ estable vs mes anterior</span>'

        # Formatear cifras
        cprev_str = f"${dist['costo_preventivo']:,.0f}".replace(",", ".")
        ccorr_str = f"${dist['costo_correctivo']:,.0f}".replace(",", ".")
        ctotal_str = f"${dist['costo_total']:,.0f}".replace(",", ".")

        # Anchos de la barra apilada (mínimo 5% para que se vean ambas si una es muy pequeña)
        ancho_p = max(5, dist['pct_preventivo']) if dist['pct_preventivo'] > 0 else 0
        ancho_c = max(5, dist['pct_correctivo']) if dist['pct_correctivo'] > 0 else 0
        # Renormalizar para sumar 100
        suma = ancho_p + ancho_c
        if suma > 0:
            ancho_p = (ancho_p / suma) * 100
            ancho_c = (ancho_c / suma) * 100

        # Lista de OTs por tipo
        otp_chips = []
        if dist['n_p']  > 0: otp_chips.append(f"<strong>{dist['n_p']}</strong> OT-P")
        if dist['n_m']  > 0: otp_chips.append(f"<strong>{dist['n_m']}</strong> OT-M")
        otc_chips = []
        if dist['n_ci'] > 0: otc_chips.append(f"<strong>{dist['n_ci']}</strong> OT-CI")
        if dist['n_co'] > 0: otc_chips.append(f"<strong>{dist['n_co']}</strong> OT-CO")
        otp_str = " · ".join(otp_chips) if otp_chips else "0 OTs"
        otc_str = " · ".join(otc_chips) if otc_chips else "0 OTs"

        # Construir HTML compactado (sin comentarios HTML, sin saltos largos
        # que confundan el parser de Markdown de Streamlit).
        html_kpi = (
            f'<div class="kpi-card {cls_card}">'
            f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:.6rem">'
            f'<div class="kpi-card-head" style="margin-bottom:0">⚖️ Distribución por costo del mes</div>'
            f'<div style="font-size:.78rem">Madurez: '
            f'<strong style="color:{color_madurez};font-size:.85rem">{dist["madurez"]}</strong>'
            f'</div></div>'
            # Barra apilada
            f'<div style="display:flex;width:100%;height:42px;border-radius:8px;overflow:hidden;'
            f'border:1px solid var(--border);box-shadow:inset 0 0 4px rgba(0,0,0,.3)">'
            f'<div style="width:{ancho_p:.1f}%;background:linear-gradient(135deg,#3FB950 0%,#5DD46F 100%);'
            f'display:flex;align-items:center;justify-content:center;'
            f'color:#0D1117;font-weight:800;font-size:1rem;font-family:\'Exo 2\',sans-serif">'
            f'{dist["pct_preventivo"]:.0f}% Preventivo</div>'
            f'<div style="width:{ancho_c:.1f}%;background:linear-gradient(135deg,#F85149 0%,#FF6B5C 100%);'
            f'display:flex;align-items:center;justify-content:center;'
            f'color:#FFFFFF;font-weight:800;font-size:1rem;font-family:\'Exo 2\',sans-serif">'
            f'{dist["pct_correctivo"]:.0f}% Correctivo</div>'
            f'</div>'
            # Detalle debajo de la barra
            f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:1rem;margin-top:.7rem">'
            f'<div style="border-left:3px solid #3FB950;padding-left:.6rem">'
            f'<div style="font-size:.7rem;color:var(--muted);text-transform:uppercase;letter-spacing:.5px">Preventivo</div>'
            f'<div style="font-family:\'Exo 2\',sans-serif;font-size:1.15rem;font-weight:700;color:var(--text)">{cprev_str}</div>'
            f'<div style="font-size:.75rem;color:var(--muted)">{otp_str}</div>'
            f'</div>'
            f'<div style="border-left:3px solid #F85149;padding-left:.6rem">'
            f'<div style="font-size:.7rem;color:var(--muted);text-transform:uppercase;letter-spacing:.5px">Correctivo</div>'
            f'<div style="font-family:\'Exo 2\',sans-serif;font-size:1.15rem;font-weight:700;color:var(--text)">{ccorr_str}</div>'
            f'<div style="font-size:.75rem;color:var(--muted)">{otc_str}</div>'
            f'</div></div>'
            # Total + tendencia
            f'<div style="margin-top:.6rem;padding-top:.5rem;border-top:1px solid var(--border);'
            f'display:flex;justify-content:space-between;align-items:center;font-size:.78rem">'
            f'<span style="color:var(--muted)">Total invertido: '
            f'<strong style="color:var(--text)">{ctotal_str}</strong></span>'
            f'{tend_html}'
            f'</div>'
            # Nota explicativa
            f'<div style="font-size:.72rem;color:var(--muted);margin-top:.5rem;line-height:1.4;font-style:italic">'
            f'ℹ️ Una distribución ≥60% preventivo indica gestión madura. '
            f'Entre 40%-60% en transición. &lt;40% modo "apaga incendios".'
            f'</div>'
            f'</div>'
        )
        st.markdown(html_kpi, unsafe_allow_html=True)

    st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════
    #  PARETO DE FALLAS
    # ═══════════════════════════════════════════════════════
    st.markdown("""<h2 style="font-family:'Exo 2',sans-serif;color:var(--teal-lt);
        font-size:1.2rem;margin:0 0 .8rem 0;letter-spacing:.5px">
        📈 Pareto de fallas correctivas — Top 10 sistemas
        </h2>""", unsafe_allow_html=True)

    pareto = calcular_pareto_fallas(df_mtto, mes_ini, mes_fin)
    if not pareto:
        st.markdown("""<div class="kpi-card kpi-neutral">
          <div class="empty-state">
            <div class="em-icon">📭</div>
            <div>No hay OTs correctivas cerradas en el mes vigente.</div>
            <div style="font-size:.78rem;margin-top:.3rem">
              El Pareto se construye con OT-CI y OT-CO cerradas durante el mes.
            </div>
          </div>
        </div>""", unsafe_allow_html=True)
    else:
        max_n = pareto[0]["n_ots"]
        n_total = len(pareto)
        # Paleta de colores por ranking (rojo → naranja → amarillo → verde)
        # Cada item recibe un color según su posición en el ranking
        def _color_por_ranking(idx: int, total: int) -> tuple:
            """Devuelve (color_inicio, color_fin, color_texto) para el degradado del bar."""
            # Posición normalizada [0..1]: 0 = peor (más fallas), 1 = mejor
            pos = idx / max(1, total - 1)
            # Definir 5 paradas: rojo intenso → rojo-naranja → naranja → amarillo → verde
            paradas = [
                (0.00, ("#F85149", "#FF6B5C", "#FFFFFF")),  # rojo intenso
                (0.25, ("#FF6B00", "#FF9540", "#000000")),  # naranja-rojo
                (0.50, ("#D29922", "#F0B83A", "#000000")),  # amarillo
                (0.75, ("#80B53F", "#A2D45A", "#000000")),  # verde-amarillo
                (1.00, ("#3FB950", "#5DD46F", "#000000")),  # verde
            ]
            # Encontrar el más cercano
            for umbral, color in paradas:
                if pos <= umbral:
                    return color
            return paradas[-1][1]

        rows_html = ""
        for idx, p in enumerate(pareto):
            ancho_pct = (p["n_ots"] / max_n) * 100 if max_n > 0 else 0
            c1, c2, ctxt = _color_por_ranking(idx, n_total)
            rows_html += f"""<div class="pareto-row">
              <div>
                <div class="pareto-label">{p['sistema']}</div>
                <div class="pareto-bar">
                  <div class="pareto-bar-fill" style="width:{ancho_pct:.0f}%;
                       background:linear-gradient(90deg, {c1} 0%, {c2} 100%);
                       color:{ctxt}">{p['n_ots']}</div>
                </div>
              </div>
              <div>
                <div class="pareto-count"><strong style="color:var(--text)">{p['pct']:.0f}%</strong></div>
                <div class="pareto-pct">acum: {p['pct_acum']:.0f}%</div>
              </div>
            </div>"""
        st.markdown(f"""<div class="kpi-card kpi-info" style="padding:.5rem 1rem">
          {rows_html}
        </div>""", unsafe_allow_html=True)
        st.markdown("""<div class="info-banner">
          ℹ️ Las barras se colorean según el ranking: rojo para los sistemas con más fallas
          (mayor prioridad de atención), degradando a verde para los menos críticos.
          El Pareto detallado por costo, vehículo, modo de falla y tendencia mensual
          está disponible en Power BI tras importar la Hoja de Vida.
        </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════
    #  ACCIONES: Descargar HV + Abrir Power BI
    # ═══════════════════════════════════════════════════════
    st.markdown("""<h2 style="font-family:'Exo 2',sans-serif;color:var(--teal-lt);
        font-size:1.2rem;margin:0 0 .8rem 0;letter-spacing:.5px">
        🔗 Análisis avanzado en Power BI
        </h2>""", unsafe_allow_html=True)

    cfg = _cargar_config()
    cb1, cb2 = st.columns([1, 1], gap="medium")

    with cb1:
        st.markdown("""<div class="kpi-card" style="border-left:5px solid var(--teal)">
          <div class="kpi-card-head">📥 Hoja de Vida Consolidada</div>
          <div style="font-size:.85rem;color:var(--text);margin:.4rem 0">
            Exporta un archivo Excel con 6 tablas listas para Power BI:
          </div>
          <ul style="font-size:.78rem;color:var(--muted);margin:0;padding-left:1.2rem;line-height:1.6">
            <li>Maestro de Vehículos</li>
            <li>Catálogo de Rutinas Preventivas</li>
            <li>Histórico de Inspecciones <span style="color:var(--teal)">(formato ancho coloreado)</span></li>
            <li>Inspecciones por Ítem <span style="color:var(--teal)">(formato largo para Power BI)</span></li>
            <li>Histórico de Mantenimientos (todas las OTs)</li>
            <li>Cumplimiento de Programación Preventiva</li>
          </ul>
        </div>""", unsafe_allow_html=True)
        try:
            xlsx_bytes = generar_hoja_de_vida()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            st.download_button(
                label="⬇️  Descargar Hoja de Vida (Excel)",
                data=xlsx_bytes,
                file_name=f"Vekmaint_HojaDeVida_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="ana_descargar_hv",
            )
        except Exception as e:
            st.error(f"Error generando el archivo: {e}")

    with cb2:
        url_actual = cfg.get("power_bi_url", "")
        st.markdown("""<div class="kpi-card" style="border-left:5px solid #F2C811">
          <div class="kpi-card-head">📊 Dashboard en Power BI</div>
          <div style="font-size:.85rem;color:var(--text);margin:.4rem 0">
            Para análisis estratégico avanzado: tendencias mensuales, drill-down por
            vehículo, comparativos interanuales y reportes para auditorías regulatorias.
          </div>
        </div>""", unsafe_allow_html=True)
        if url_actual:
            st.link_button("🚀  Abrir Dashboard en Power BI",
                            url=url_actual, use_container_width=True)
        else:
            st.markdown("""<div style="padding:.6rem .8rem;background:rgba(255,107,0,.08);
              border:1px solid rgba(255,107,0,.3);border-radius:8px;font-size:.78rem;
              color:#FF9A40;margin-top:.3rem">
              ⚙️ Power BI aún no está enlazado. Configure la URL abajo.
            </div>""", unsafe_allow_html=True)

        with st.expander("⚙️ Configurar URL del Power BI", expanded=not bool(url_actual)):
            st.caption("Pegue la URL pública del reporte o dashboard publicado en el servicio Power BI.")
            url_nueva = st.text_input("URL del reporte Power BI",
                                         value=url_actual,
                                         placeholder="https://app.powerbi.com/view?...",
                                         key="ana_url_pbi")
            cg1, cg2 = st.columns(2)
            with cg1:
                if st.button("💾 Guardar URL", use_container_width=True, key="ana_guardar_url"):
                    cfg["power_bi_url"] = url_nueva.strip()
                    _guardar_config(cfg)
                    st.success("✓ URL guardada")
                    st.rerun()
            with cg2:
                if url_actual and st.button("🗑 Eliminar URL", use_container_width=True, key="ana_eliminar_url"):
                    cfg["power_bi_url"] = ""
                    _guardar_config(cfg)
                    st.success("✓ URL eliminada")
                    st.rerun()


if __name__ == "__main__":
    run()
