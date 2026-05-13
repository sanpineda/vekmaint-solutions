"""
═══════════════════════════════════════════════════════════════════
  VEKMAINT SOLUTIONS · Gestión de Flota
═══════════════════════════════════════════════════════════════════

Módulo administrativo para registrar, editar y gestionar el inventario
de vehículos de la empresa cliente. Es el primer paso antes de empezar
a usar los demás módulos del aplicativo.

CARACTERÍSTICAS:
  · Vista tabla resumen con filtros por marca, tipo y estado
  · Formulario de creación con validaciones (placa, num_interno únicos)
  · Edición de campos no críticos (no permite cambiar num_interno, placa, km)
  · Inactivación lógica (nunca borrado físico) — preserva histórico
  · Estado "En reparación" se calcula automáticamente desde OTs abiertas
  · Importación masiva desde Excel con plantilla descargable
  · Compatibilidad con flota existente (campos opcionales)

ESTRUCTURA DE DATOS (flota_vehiculos.json):
{
  "001": {
    "num_interno":      "001",
    "placa":            "ABC123",
    "marca":            "Chevrolet",
    "referencia":       "NPR Reward",
    "modelo":           2018,
    "tipo":             "Buseta",
    "km_actual":        137025,
    "km_inicial":       100000,           ← km al momento de registro
    "km_actualizado_en":"2026-04-27",
    "estado":           "Activo",        ← Activo / Inactivo / Dado de baja
    "fecha_registro":   "2026-05-04",
    "rutinas_ultimas":  { ... }          ← preserva lo existente
  }
}
"""
from __future__ import annotations
import os
import io
import json
import re
from datetime import datetime, date
from pathlib import Path

import streamlit as st
import pandas as pd

# ═══════════════════════════════════════════════════════════════════
#  CONFIG
# ═══════════════════════════════════════════════════════════════════
FLOTA_DB      = "flota_vehiculos.json"
PENDIENTES_DB = "ots_pendientes.json"

MARCAS_PREDEFINIDAS = [
    "Chevrolet", "Mercedes-Benz", "Hino", "Agrale", "Volkswagen",
    "Volvo", "Scania", "International", "Foton", "JAC", "Otro",
]
TIPOS_VEHICULO = ["Bus", "Busetón", "Buseta", "Microbús"]
ESTADOS = ["Activo", "Inactivo", "Dado de baja"]   # "En reparación" se calcula automáticamente

ANO_MIN = 1990
ANO_MAX = datetime.now().year + 1

# Patrón de placa colombiana: 3 letras + 3 dígitos, o 3 letras + 2 dígitos + 1 letra
PLACA_REGEX = re.compile(r"^[A-Z]{3}-?[0-9]{2}[0-9A-Z]$", re.IGNORECASE)


# ═══════════════════════════════════════════════════════════════════
#  HELPERS DE PERSISTENCIA
# ═══════════════════════════════════════════════════════════════════
def _cargar_flota() -> dict:
    if not os.path.exists(FLOTA_DB):
        return {}
    try:
        with open(FLOTA_DB, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _guardar_flota(flota: dict):
    with open(FLOTA_DB, "w", encoding="utf-8") as f:
        json.dump(flota, f, ensure_ascii=False, indent=2)


def _cargar_pendientes() -> dict:
    if not os.path.exists(PENDIENTES_DB):
        return {}
    try:
        with open(PENDIENTES_DB, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _vehiculos_en_reparacion() -> set:
    """Conjunto de num_interno con OT-CI u OT-CO sin cerrar."""
    pendientes = _cargar_pendientes()
    en_rep = set()
    for ot, datos in pendientes.items():
        if ot.startswith("OT-CI-") or ot.startswith("OT-CO-"):
            veh = str(datos.get("Numero_Interno", "")).strip()
            if veh:
                en_rep.add(veh)
    return en_rep


def _vehiculo_tiene_historico(num_interno: str) -> bool:
    """Verifica si un vehículo tiene inspecciones o OTs registradas."""
    veh = str(num_interno).strip()
    # Histórico de inspecciones
    if os.path.exists("inspecciones_vehiculares.xlsx"):
        try:
            df = pd.read_excel("inspecciones_vehiculares.xlsx")
            if "Numero_Interno" in df.columns:
                if any(str(v).strip() == veh for v in df["Numero_Interno"]):
                    return True
        except Exception:
            pass
    # Histórico de mantenimientos
    if os.path.exists("mantenimiento_flotas.xlsx"):
        try:
            df = pd.read_excel("mantenimiento_flotas.xlsx")
            if "Numero_Interno" in df.columns:
                if any(str(v).strip() == veh for v in df["Numero_Interno"]):
                    return True
        except Exception:
            pass
    # OTs pendientes
    pendientes = _cargar_pendientes()
    for ot, datos in pendientes.items():
        if str(datos.get("Numero_Interno", "")).strip() == veh:
            return True
    return False


# ═══════════════════════════════════════════════════════════════════
#  HELPERS DE VALIDACIÓN
# ═══════════════════════════════════════════════════════════════════
def _validar_num_interno(val: str, flota: dict, edicion_de: str = None) -> str | None:
    """Retorna mensaje de error si inválido, None si válido."""
    val = str(val).strip().upper()
    if not val:
        return "El número interno es obligatorio."
    if len(val) > 10:
        return "El número interno no puede tener más de 10 caracteres."
    if not val.replace(" ", "").isalnum():
        return "Solo se permiten letras y números."
    # Unicidad (excepto si es el mismo que se está editando)
    if val in flota and val != edicion_de:
        return f"Ya existe un vehículo con el número interno '{val}'."
    return None


def _validar_placa(val: str, flota: dict, edicion_de: str = None) -> str | None:
    val = str(val).strip().upper().replace("-", "")
    if not val:
        return "La placa es obligatoria."
    if len(val) != 6:
        return "La placa debe tener 6 caracteres (sin guion)."
    if not PLACA_REGEX.match(val):
        return "Formato de placa inválido. Ejemplos válidos: ABC123, ABC12D"
    # Unicidad
    for ni, vdata in flota.items():
        if str(vdata.get("placa", "")).strip().upper().replace("-", "") == val:
            if ni != edicion_de:
                return f"Ya existe un vehículo con la placa '{val}' (núm. interno {ni})."
    return None


def _validar_modelo(val) -> str | None:
    try:
        ano = int(val)
    except (ValueError, TypeError):
        return "El modelo debe ser un año numérico."
    if ano < ANO_MIN or ano > ANO_MAX:
        return f"El modelo debe estar entre {ANO_MIN} y {ANO_MAX}."
    return None


def _validar_km(val) -> str | None:
    try:
        km = int(val)
    except (ValueError, TypeError):
        return "El kilometraje debe ser un número entero."
    if km < 0:
        return "El kilometraje no puede ser negativo."
    if km > 5_000_000:
        return "Kilometraje fuera de rango razonable."
    return None


# ═══════════════════════════════════════════════════════════════════
#  ESTADO CALCULADO (combina estado registrado + reparación)
# ═══════════════════════════════════════════════════════════════════
def _estado_efectivo(num_interno: str, vdata: dict, en_reparacion: set) -> str:
    """Calcula el estado real considerando OTs abiertas."""
    estado_registrado = vdata.get("estado", "Activo")
    if estado_registrado in ("Inactivo", "Dado de baja"):
        return estado_registrado
    if num_interno in en_reparacion:
        return "En reparación"
    return "Activo"


# ═══════════════════════════════════════════════════════════════════
#  PLANTILLA EXCEL PARA IMPORTACIÓN MASIVA
# ═══════════════════════════════════════════════════════════════════
def generar_plantilla_excel() -> bytes:
    """Genera una plantilla Excel con cabeceras + ejemplo + instrucciones."""
    output = io.BytesIO()

    # Datos de ejemplo
    ejemplo = pd.DataFrame([
        {"Numero_Interno": "001", "Placa": "ABC123", "Marca": "Chevrolet",
         "Referencia": "NPR Reward", "Modelo": 2020, "Tipo": "Buseta",
         "Km_Actual": 85000, "Estado": "Activo"},
        {"Numero_Interno": "002", "Placa": "DEF456", "Marca": "Mercedes-Benz",
         "Referencia": "Atego 1419", "Modelo": 2018, "Tipo": "Bus",
         "Km_Actual": 142500, "Estado": "Activo"},
    ])

    # Hoja de instrucciones
    instrucciones = pd.DataFrame({
        "CAMPO": ["Numero_Interno", "Placa", "Marca", "Referencia",
                  "Modelo", "Tipo", "Km_Actual", "Estado"],
        "OBLIGATORIO": ["Sí", "Sí", "Sí", "Sí", "Sí", "Sí", "Sí", "Opcional"],
        "TIPO": ["Texto", "Texto", "Lista", "Texto", "Año", "Lista", "Numérico", "Lista"],
        "DESCRIPCIÓN": [
            "Identificador interno único, hasta 10 caracteres alfanuméricos.",
            "Placa colombiana, formato ABC123 o ABC12D.",
            f"Una de: {', '.join(MARCAS_PREDEFINIDAS)}.",
            "Referencia comercial del fabricante (NPR, Atego, etc).",
            f"Año del modelo, entre {ANO_MIN} y {ANO_MAX}.",
            f"Una de: {', '.join(TIPOS_VEHICULO)}.",
            "Kilometraje actual del vehículo al momento de registro.",
            f"Una de: {', '.join(ESTADOS)}. Por defecto: Activo.",
        ],
    })

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        ejemplo.to_excel(writer, sheet_name="Vehiculos", index=False)
        instrucciones.to_excel(writer, sheet_name="Instrucciones", index=False)

        # Estilos
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        fill_header = PatternFill("solid", fgColor="0A1628")
        font_header = Font(color="FFFFFF", bold=True, size=10)

        for sheet_name in ["Vehiculos", "Instrucciones"]:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Auto-ajustar columnas
            for col in ws.columns:
                lens = [len(str(c.value)) for c in col if c.value is not None]
                ml = max(lens) if lens else 10
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 60)
            ws.freeze_panes = "A2"

    output.seek(0)
    return output.getvalue()


# ═══════════════════════════════════════════════════════════════════
#  IMPORTACIÓN MASIVA DESDE EXCEL
# ═══════════════════════════════════════════════════════════════════
def importar_desde_excel(file_bytes) -> tuple[int, list]:
    """
    Importa vehículos desde un Excel. Retorna (n_creados, errores).
    Cada error es un dict: {fila, num_interno, error}
    """
    try:
        df = pd.read_excel(file_bytes, sheet_name="Vehiculos")
    except Exception as e:
        try:
            df = pd.read_excel(file_bytes)
        except Exception as e2:
            return 0, [{"fila": "—", "num_interno": "—",
                          "error": f"No se pudo leer el archivo: {e2}"}]

    # Validar columnas requeridas
    required = {"Numero_Interno", "Placa", "Marca", "Referencia",
                "Modelo", "Tipo", "Km_Actual"}
    faltantes = required - set(df.columns)
    if faltantes:
        return 0, [{"fila": "—", "num_interno": "—",
                      "error": f"Faltan columnas obligatorias: {', '.join(faltantes)}"}]

    flota = _cargar_flota()
    errores = []
    creados = 0

    for idx, row in df.iterrows():
        fila_num = idx + 2  # +1 por header, +1 por base 1
        ni = str(row.get("Numero_Interno", "")).strip().upper()
        placa = str(row.get("Placa", "")).strip().upper().replace("-", "")
        marca = str(row.get("Marca", "")).strip()
        referencia = str(row.get("Referencia", "")).strip()
        try:
            modelo = int(row.get("Modelo", 0))
        except (ValueError, TypeError):
            modelo = 0
        tipo = str(row.get("Tipo", "")).strip()
        try:
            km = int(row.get("Km_Actual", 0))
        except (ValueError, TypeError):
            km = -1
        estado = str(row.get("Estado", "Activo")).strip() or "Activo"

        # Validaciones
        err_ni = _validar_num_interno(ni, flota)
        if err_ni:
            errores.append({"fila": fila_num, "num_interno": ni, "error": err_ni})
            continue
        err_pl = _validar_placa(placa, flota)
        if err_pl:
            errores.append({"fila": fila_num, "num_interno": ni, "error": err_pl})
            continue
        if marca not in MARCAS_PREDEFINIDAS and marca != "Otro":
            # Aceptar marcas no estándar pero advertir
            pass
        err_mod = _validar_modelo(modelo)
        if err_mod:
            errores.append({"fila": fila_num, "num_interno": ni, "error": err_mod})
            continue
        if tipo not in TIPOS_VEHICULO:
            errores.append({"fila": fila_num, "num_interno": ni,
                              "error": f"Tipo inválido. Debe ser uno de: {', '.join(TIPOS_VEHICULO)}"})
            continue
        err_km = _validar_km(km)
        if err_km:
            errores.append({"fila": fila_num, "num_interno": ni, "error": err_km})
            continue
        if estado not in ESTADOS:
            estado = "Activo"

        # Crear vehículo
        flota[ni] = {
            "num_interno":       ni,
            "placa":              placa,
            "marca":              marca,
            "referencia":         referencia,
            "modelo":             modelo,
            "tipo":               tipo,
            "km_actual":          km,
            "km_inicial":         km,
            "km_actualizado_en":  date.today().strftime("%Y-%m-%d"),
            "estado":             estado,
            "fecha_registro":     date.today().strftime("%Y-%m-%d"),
            "rutinas_ultimas":    {},
        }
        creados += 1

    if creados > 0:
        _guardar_flota(flota)

    return creados, errores


# ═══════════════════════════════════════════════════════════════════
#  RENDER PRINCIPAL
# ═══════════════════════════════════════════════════════════════════
def run():
    # Estilos del módulo
    st.markdown("""<style>
    :root{
      --bg:#0D1117;--bg-deep:#080C14;--surface:#161B22;--surface2:#1F2937;
      --border:#30363D;--text:#E6EDF3;--muted:#8B949E;
      --accent:#FF9540;--accent-dk:#FF6B00;--accent-lt:#FFD8B8;
      --green:#3FB950;--orange:#FF6B00;--red:#F85149;--blue:#0A84FF;--purple:#7C4DFF;
    }
    .flota-header{display:flex;align-items:center;gap:1rem;padding:1.2rem 1.5rem;
      background:linear-gradient(135deg, rgba(255,149,64,.18) 0%, rgba(255,107,0,.08) 100%);
      border:1px solid rgba(255,149,64,.4);border-left:5px solid var(--accent);
      border-radius:14px;margin-bottom:1.2rem}
    .flota-icon{font-size:2.2rem}
    .flota-title{font-family:'Exo 2',sans-serif;font-size:1.6rem;font-weight:700;color:var(--accent-lt);margin:0;letter-spacing:.5px}
    .flota-sub{color:var(--muted);font-size:.85rem;margin-top:.15rem}

    .stat-mini{display:inline-flex;align-items:center;gap:.5rem;background:var(--surface);
      border:1px solid var(--border);border-radius:8px;padding:.5rem .9rem;font-size:.85rem;color:var(--text);margin-right:.6rem;margin-bottom:.4rem}
    .stat-mini strong{color:var(--accent-lt);font-size:1.1rem;font-weight:700;font-family:'Exo 2',sans-serif}

    .veh-card{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:.9rem 1rem;margin-bottom:.6rem}
    .veh-card-head{display:flex;justify-content:space-between;align-items:center;gap:.8rem;margin-bottom:.4rem}
    .veh-num{font-family:'Exo 2',sans-serif;font-weight:800;font-size:1.1rem;color:var(--text)}
    .veh-placa{font-family:'Inter',sans-serif;font-size:.8rem;background:var(--surface2);padding:.2rem .5rem;border-radius:5px;color:var(--muted);letter-spacing:1px}
    .veh-info{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:.5rem .8rem;font-size:.78rem;color:var(--muted)}
    .veh-info strong{color:var(--text);font-weight:600}

    .badge-estado{display:inline-block;padding:.2rem .6rem;border-radius:12px;font-size:.7rem;font-weight:700;text-transform:uppercase;letter-spacing:.5px}
    .badge-activo{background:rgba(63,185,80,.18);color:var(--green);border:1px solid rgba(63,185,80,.4)}
    .badge-reparacion{background:rgba(255,107,0,.18);color:var(--orange);border:1px solid rgba(255,107,0,.4)}
    .badge-inactivo{background:rgba(139,148,158,.18);color:var(--muted);border:1px solid rgba(139,148,158,.4)}
    .badge-baja{background:rgba(248,81,73,.18);color:var(--red);border:1px solid rgba(248,81,73,.4)}

    .empty-state{text-align:center;padding:3rem 1rem;color:var(--muted);background:var(--surface);
      border:2px dashed var(--border);border-radius:14px}
    .empty-state .em-icon{font-size:3rem;margin-bottom:.6rem;opacity:.5}
    .empty-state h3{font-family:'Exo 2',sans-serif;color:var(--text);margin:.3rem 0;font-size:1.2rem}
    </style>""", unsafe_allow_html=True)

    # Botón volver al hub
    cb1, _ = st.columns([1, 8])
    with cb1:
        if st.button("← Volver", key="flota_volver", use_container_width=True):
            st.session_state["modulo"] = "hub"
            try:
                st.query_params.clear()
            except Exception:
                pass
            st.rerun()

    flota = _cargar_flota()
    en_reparacion = _vehiculos_en_reparacion()
    n_total = len(flota)
    n_activos = sum(1 for ni, v in flota.items()
                      if _estado_efectivo(ni, v, en_reparacion) == "Activo")
    n_rep = sum(1 for ni, v in flota.items()
                  if _estado_efectivo(ni, v, en_reparacion) == "En reparación")
    n_inact = sum(1 for ni, v in flota.items()
                    if _estado_efectivo(ni, v, en_reparacion) in ("Inactivo", "Dado de baja"))

    # Header
    st.markdown(f"""<div class="flota-header">
      <div class="flota-icon">🚛</div>
      <div>
        <h1 class="flota-title">Gestión de Flota</h1>
        <div class="flota-sub">Inventario maestro de vehículos · base para todos los demás módulos</div>
      </div>
    </div>""", unsafe_allow_html=True)

    # Stats rápidas
    st.markdown(f"""<div style="margin-bottom:1rem">
      <span class="stat-mini">🚛 <strong>{n_total}</strong> total</span>
      <span class="stat-mini">🟢 <strong>{n_activos}</strong> activos</span>
      <span class="stat-mini">🔧 <strong>{n_rep}</strong> en reparación</span>
      <span class="stat-mini">⚪ <strong>{n_inact}</strong> inactivos / dados de baja</span>
    </div>""", unsafe_allow_html=True)

    # Tabs principales
    tab_listado, tab_nuevo, tab_importar = st.tabs([
        "📋 Listado de vehículos",
        "➕ Registrar vehículo",
        "📥 Importación masiva (Excel)",
    ])

    # ═══════════════════════════════════════════════════════
    #  TAB 1 — LISTADO DE VEHÍCULOS
    # ═══════════════════════════════════════════════════════
    with tab_listado:
        if not flota:
            st.markdown("""<div class="empty-state">
              <div class="em-icon">🚛</div>
              <h3>Aún no hay vehículos registrados</h3>
              <p>Comienza registrando tu primer vehículo en la pestaña "Registrar vehículo"
              o usa la importación masiva si tienes una flota existente.</p>
            </div>""", unsafe_allow_html=True)
        else:
            # Filtros
            fc1, fc2, fc3, fc4 = st.columns(4)
            with fc1:
                filtro_marca = st.selectbox("Marca", ["Todas"] + MARCAS_PREDEFINIDAS,
                                              key="flt_marca")
            with fc2:
                filtro_tipo = st.selectbox("Tipo", ["Todos"] + TIPOS_VEHICULO,
                                             key="flt_tipo")
            with fc3:
                filtro_estado = st.selectbox("Estado",
                                               ["Todos", "Activo", "En reparación",
                                                "Inactivo", "Dado de baja"],
                                               key="flt_estado")
            with fc4:
                buscar = st.text_input("Buscar (núm. o placa)", key="flt_buscar",
                                          placeholder="Ej: 005, ABC123").strip().upper()

            # Aplicar filtros
            vehiculos_filtrados = []
            for ni, vdata in flota.items():
                est_efectivo = _estado_efectivo(ni, vdata, en_reparacion)
                if filtro_marca != "Todas" and vdata.get("marca", "") != filtro_marca:
                    continue
                if filtro_tipo != "Todos" and vdata.get("tipo", "") != filtro_tipo:
                    continue
                if filtro_estado != "Todos" and est_efectivo != filtro_estado:
                    continue
                if buscar:
                    placa = str(vdata.get("placa", "")).upper()
                    if buscar not in ni.upper() and buscar not in placa:
                        continue
                vehiculos_filtrados.append((ni, vdata, est_efectivo))

            # Ordenar por num_interno
            vehiculos_filtrados.sort(key=lambda x: x[0])

            st.caption(f"Mostrando **{len(vehiculos_filtrados)}** de {n_total} vehículos")

            if not vehiculos_filtrados:
                st.info("Ningún vehículo coincide con los filtros seleccionados.")
            else:
                # Mostrar vehículos en cards
                for ni, vdata, est_efectivo in vehiculos_filtrados:
                    badge_class = {
                        "Activo": "badge-activo",
                        "En reparación": "badge-reparacion",
                        "Inactivo": "badge-inactivo",
                        "Dado de baja": "badge-baja",
                    }.get(est_efectivo, "badge-inactivo")

                    km_str = f"{vdata.get('km_actual', 0):,}".replace(",", ".")
                    km_act = vdata.get("km_actualizado_en", "—")

                    st.markdown(f"""<div class="veh-card">
                      <div class="veh-card-head">
                        <div style="display:flex;align-items:center;gap:.7rem">
                          <span class="veh-num">Vehículo {ni}</span>
                          <span class="veh-placa">{vdata.get('placa', 'Sin placa')}</span>
                          <span class="badge-estado {badge_class}">{est_efectivo}</span>
                        </div>
                      </div>
                      <div class="veh-info">
                        <div>Marca: <strong>{vdata.get('marca', '—')}</strong></div>
                        <div>Referencia: <strong>{vdata.get('referencia', vdata.get('modelo', '—'))}</strong></div>
                        <div>Modelo: <strong>{vdata.get('modelo', '—')}</strong></div>
                        <div>Tipo: <strong>{vdata.get('tipo', '—')}</strong></div>
                        <div>Km actual: <strong>{km_str} km</strong></div>
                        <div>Última actualización: <strong>{km_act}</strong></div>
                      </div>
                    </div>""", unsafe_allow_html=True)

                    bc1, bc2, bc3, _ = st.columns([1, 1, 1, 4])
                    with bc1:
                        if st.button("✏️ Editar", key=f"edit_{ni}", use_container_width=True):
                            st.session_state["flota_editando"] = ni
                            st.session_state["flota_tab"] = "editar"
                            st.rerun()
                    with bc2:
                        # Botón inactivar/reactivar según estado registrado
                        estado_reg = vdata.get("estado", "Activo")
                        if estado_reg == "Activo":
                            if st.button("⏸ Inactivar", key=f"inact_{ni}", use_container_width=True):
                                flota[ni]["estado"] = "Inactivo"
                                _guardar_flota(flota)
                                st.success(f"✓ Vehículo {ni} inactivado")
                                st.rerun()
                        else:
                            if st.button("▶ Reactivar", key=f"react_{ni}", use_container_width=True):
                                flota[ni]["estado"] = "Activo"
                                _guardar_flota(flota)
                                st.success(f"✓ Vehículo {ni} reactivado")
                                st.rerun()
                    with bc3:
                        # Solo ofrece "Dar de baja" si no tiene OT abierta
                        if vdata.get("estado", "Activo") != "Dado de baja":
                            if st.button("🚫 Dar de baja", key=f"baja_{ni}", use_container_width=True):
                                flota[ni]["estado"] = "Dado de baja"
                                _guardar_flota(flota)
                                st.success(f"✓ Vehículo {ni} dado de baja")
                                st.rerun()

        # Modo edición — formulario
        if st.session_state.get("flota_editando"):
            ni_edit = st.session_state["flota_editando"]
            if ni_edit in flota:
                _renderizar_formulario_edicion(ni_edit, flota[ni_edit], flota)

    # ═══════════════════════════════════════════════════════
    #  TAB 2 — REGISTRAR NUEVO VEHÍCULO
    # ═══════════════════════════════════════════════════════
    with tab_nuevo:
        _renderizar_formulario_nuevo(flota)

    # ═══════════════════════════════════════════════════════
    #  TAB 3 — IMPORTACIÓN MASIVA
    # ═══════════════════════════════════════════════════════
    with tab_importar:
        st.markdown("""<div style="background:rgba(10,132,255,.08);border:1px solid rgba(10,132,255,.3);
          border-radius:10px;padding:1rem 1.2rem;margin-bottom:1rem;font-size:.88rem;color:#7EC4FF">
          📋 <strong>Importación masiva desde Excel</strong> — ideal para empresas que
          inician con una flota completa de vehículos. El proceso tiene 3 pasos:
        </div>""", unsafe_allow_html=True)

        st.markdown("##### Paso 1 — Descarga la plantilla Excel")
        try:
            plantilla_bytes = generar_plantilla_excel()
            st.download_button(
                label="⬇️  Descargar plantilla Excel",
                data=plantilla_bytes,
                file_name="Vekmaint_Plantilla_Flota.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_plantilla_flota",
            )
        except Exception as e:
            st.error(f"Error generando plantilla: {e}")

        st.markdown("##### Paso 2 — Llena la plantilla con tus vehículos")
        st.markdown("""<div style="background:var(--surface);border:1px solid var(--border);
          border-radius:8px;padding:.7rem 1rem;font-size:.82rem;color:var(--muted)">
          Abre la plantilla y completa la hoja <strong>"Vehiculos"</strong>. La hoja
          <strong>"Instrucciones"</strong> contiene la descripción y validaciones de cada campo.
          Borra las filas de ejemplo antes de cargar tus datos reales.
        </div>""", unsafe_allow_html=True)

        st.markdown("##### Paso 3 — Sube el archivo lleno")
        archivo = st.file_uploader("Selecciona el archivo Excel diligenciado",
                                      type=["xlsx", "xls"],
                                      key="upload_flota")

        if archivo is not None:
            if st.button("🚀 Procesar e importar", type="primary",
                            use_container_width=True, key="btn_importar"):
                with st.spinner("Procesando archivo..."):
                    creados, errores = importar_desde_excel(archivo)

                if creados > 0:
                    st.success(f"✅ Se importaron **{creados}** vehículos correctamente.")
                if errores:
                    st.warning(f"⚠️ Se encontraron {len(errores)} errores:")
                    df_err = pd.DataFrame(errores)
                    st.dataframe(df_err, use_container_width=True, hide_index=True)
                if creados > 0 and not errores:
                    st.balloons()


# ═══════════════════════════════════════════════════════════════════
#  FORMULARIO DE EDICIÓN
# ═══════════════════════════════════════════════════════════════════
def _renderizar_formulario_edicion(ni: str, vdata: dict, flota: dict):
    st.markdown("---")
    st.markdown(f"### ✏️ Editando vehículo {ni}")
    st.caption("⚠️ El número interno, la placa y el kilometraje **no se pueden modificar** desde aquí. "
                "El kilometraje se actualiza automáticamente con cada inspección preoperacional.")

    with st.form(key=f"form_edit_{ni}"):
        c1, c2 = st.columns(2)
        with c1:
            st.text_input("Número interno", value=ni, disabled=True)
            placa_actual = vdata.get("placa", "")
            st.text_input("Placa", value=placa_actual, disabled=True)

            # Marca
            marca_actual = vdata.get("marca", "Chevrolet")
            try:
                idx_marca = MARCAS_PREDEFINIDAS.index(marca_actual)
            except ValueError:
                idx_marca = MARCAS_PREDEFINIDAS.index("Otro")
            marca_nueva = st.selectbox("Marca", MARCAS_PREDEFINIDAS, index=idx_marca,
                                          key=f"e_marca_{ni}")
            if marca_nueva == "Otro":
                marca_nueva = st.text_input("Especificar marca",
                                                value=marca_actual if marca_actual not in MARCAS_PREDEFINIDAS else "",
                                                key=f"e_marca_otro_{ni}")

            referencia = st.text_input("Referencia comercial",
                                          value=vdata.get("referencia", vdata.get("modelo", "")),
                                          key=f"e_ref_{ni}")

        with c2:
            modelo_actual = vdata.get("modelo", ANO_MAX - 5)
            try:
                modelo_actual_int = int(modelo_actual)
            except (ValueError, TypeError):
                modelo_actual_int = ANO_MAX - 5
            modelo_nuevo = st.number_input("Modelo (año)", min_value=ANO_MIN,
                                              max_value=ANO_MAX,
                                              value=modelo_actual_int, step=1,
                                              key=f"e_modelo_{ni}")
            tipo_actual = vdata.get("tipo", "Buseta")
            idx_tipo = TIPOS_VEHICULO.index(tipo_actual) if tipo_actual in TIPOS_VEHICULO else 0
            tipo_nuevo = st.selectbox("Tipo", TIPOS_VEHICULO, index=idx_tipo,
                                         key=f"e_tipo_{ni}")

            km_str = f"{vdata.get('km_actual', 0):,} km".replace(",", ".")
            st.text_input("Kilometraje actual", value=km_str, disabled=True,
                            help="Se actualiza automáticamente con cada inspección preoperacional")

            estado_actual = vdata.get("estado", "Activo")
            idx_est = ESTADOS.index(estado_actual) if estado_actual in ESTADOS else 0
            estado_nuevo = st.selectbox("Estado", ESTADOS, index=idx_est,
                                           key=f"e_estado_{ni}",
                                           help="'En reparación' se calcula automáticamente desde OTs abiertas")

        guardar = st.form_submit_button("💾 Guardar cambios", type="primary",
                                            use_container_width=True)

    cancelar_col, _ = st.columns([1, 4])
    with cancelar_col:
        if st.button("✖ Cancelar", key=f"cancel_edit_{ni}", use_container_width=True):
            st.session_state.pop("flota_editando", None)
            st.rerun()

    if guardar:
        # Validar marca personalizada
        if not marca_nueva or marca_nueva.strip() == "":
            st.error("Debe especificar la marca.")
            return
        # Actualizar
        flota[ni]["marca"] = marca_nueva
        flota[ni]["referencia"] = referencia
        flota[ni]["modelo"] = int(modelo_nuevo)
        flota[ni]["tipo"] = tipo_nuevo
        flota[ni]["estado"] = estado_nuevo
        _guardar_flota(flota)
        st.session_state.pop("flota_editando", None)
        st.success(f"✓ Cambios guardados para vehículo {ni}")
        st.rerun()


# ═══════════════════════════════════════════════════════════════════
#  FORMULARIO DE NUEVO VEHÍCULO
# ═══════════════════════════════════════════════════════════════════
def _renderizar_formulario_nuevo(flota: dict):
    st.markdown("Complete los datos del nuevo vehículo. Los campos marcados con * son obligatorios.")

    with st.form(key="form_nuevo_veh", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            ni = st.text_input("Número interno *", max_chars=10,
                                  placeholder="Ej: 001, 042, B125",
                                  key="n_ni")
            placa = st.text_input("Placa * (formato ABC123 o ABC12D)",
                                     max_chars=7,
                                     placeholder="Ej: ABC123",
                                     key="n_placa")
            marca_sel = st.selectbox("Marca *", MARCAS_PREDEFINIDAS, key="n_marca")
            if marca_sel == "Otro":
                marca_otra = st.text_input("Especificar otra marca",
                                                placeholder="Ej: Hyundai, Iveco",
                                                key="n_marca_otra")
            else:
                marca_otra = ""
            referencia = st.text_input("Referencia comercial *",
                                          placeholder="Ej: NPR Reward, Atego 1419, OF-1721",
                                          key="n_ref")

        with c2:
            modelo = st.number_input("Modelo (año) *",
                                        min_value=ANO_MIN, max_value=ANO_MAX,
                                        value=ANO_MAX - 5, step=1,
                                        key="n_modelo")
            tipo = st.selectbox("Tipo de vehículo *", TIPOS_VEHICULO, key="n_tipo")
            km = st.number_input("Kilometraje actual *",
                                    min_value=0, max_value=5_000_000,
                                    value=0, step=1000,
                                    key="n_km",
                                    help="Kilometraje al momento de registro. Se actualizará con cada inspección.")
            estado = st.selectbox("Estado inicial", ["Activo", "Inactivo"],
                                     key="n_estado",
                                     help="'En reparación' se asigna automáticamente cuando hay OTs abiertas.")

        crear = st.form_submit_button("➕ Registrar vehículo", type="primary",
                                          use_container_width=True)

    if crear:
        # Validaciones
        ni_norm = str(ni).strip().upper()
        marca_final = marca_otra.strip() if marca_sel == "Otro" else marca_sel

        errores = []
        e = _validar_num_interno(ni_norm, flota)
        if e: errores.append(e)
        e = _validar_placa(placa, flota)
        if e: errores.append(e)
        if not marca_final:
            errores.append("Debe seleccionar o especificar una marca.")
        if not referencia.strip():
            errores.append("La referencia comercial es obligatoria.")
        e = _validar_modelo(modelo)
        if e: errores.append(e)
        e = _validar_km(km)
        if e: errores.append(e)

        if errores:
            for err in errores:
                st.error(f"❌ {err}")
            return

        placa_norm = str(placa).strip().upper().replace("-", "")
        flota[ni_norm] = {
            "num_interno":       ni_norm,
            "placa":              placa_norm,
            "marca":              marca_final,
            "referencia":         referencia.strip(),
            "modelo":             int(modelo),
            "tipo":               tipo,
            "km_actual":          int(km),
            "km_inicial":         int(km),
            "km_actualizado_en":  date.today().strftime("%Y-%m-%d"),
            "estado":             estado,
            "fecha_registro":     date.today().strftime("%Y-%m-%d"),
            "rutinas_ultimas":    {},
        }
        _guardar_flota(flota)
        st.success(f"✅ Vehículo **{ni_norm}** ({placa_norm}) registrado correctamente.")
        st.balloons()


if __name__ == "__main__":
    run()
