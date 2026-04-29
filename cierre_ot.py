"""
cierre_ot.py — v1.0  (antes mantenimiento_flotas.py)
=======================================================
Módulo EXCLUSIVAMENTE para cerrar órdenes de trabajo.
Standalone : streamlit run cierre_ot.py
Desde hub  : import cierre_ot; cierre_ot.run()

Cambios vs versión anterior:
----------------------------
• Eliminada la Fase 1 (reporte) — ahora vive en reporte_fallas.py y preventivo
• Este módulo solo carga una OT pendiente y la cierra técnicamente
• Soporta todos los prefijos:
    OT-CI-*  : correctivo desde inspección preoperacional
    OT-CO-*  : correctivo operación (del módulo reporte de fallas)
    OT-M-*   : correctivo mayor
    OT-P-*   : preventivo (del módulo preventivo) — pre-carga rutinas y costos
• Causa raíz: solo visible y obligatoria para correctivos (CI, CO, M)
• Cierre preventivo: valida ejecución de rutinas y actualiza flota_vehiculos.json
• Todos los campos incluyen Criticidad heredada del reporte original
"""
import streamlit as st
import pandas as pd
import json
import os
from datetime import datetime
from pathlib import Path

# ═══════════════════════════════════════════════════════════════════
#  CONSTANTES
# ═══════════════════════════════════════════════════════════════════
TIPOS_MANT = {"P": "Preventivo", "C": "Correctivo", "M": "Mayor"}
TIPOS_CORRECTIVOS = {"C", "M"}  # los que usan causa_raiz

EXCEL_FILE          = "mantenimiento_flotas.xlsx"
PENDIENTES_DB       = "ots_pendientes.json"
FLOTA_DB            = "flota_vehiculos.json"
SOPORTES_DIR        = Path("mantenimiento_soportes")

COLUMNAS_EXCEL = [
    "Fecha_Inicio_Inactividad", "Fecha_Registro_F1", "Fecha_Cierre_F2",
    "OT", "Numero_Interno", "Tipo_Mantenimiento", "Estado_OT",
    "Criticidad", "Kilometraje", "Sistema", "Modo_Falla", "Causa_Raiz",
    "Descripcion", "Cantidad", "Costo_Repuestos",
    "Costo_Mano_Obra", "Costo_Total_OT",
    "Costo_Estimado_Total",          # Solo preventivo: diferencia estimado vs real
    "Proveedor", "Tiempo_Repuesto_Dias", "Tiempo_Repuesto_Horas",
    "Responsable_Tecnico", "Conductor", "Soporte", "Origen",
]

# ═══════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════
def _fmt_moneda(v):
    return f"$ {v:,.0f}".replace(",", ".") if v else "$ 0"


def _info_vehiculo(veh_id):
    """Retorna info de la flota (marca, modelo, km_actual) si el vehículo existe."""
    if not veh_id or not os.path.exists(FLOTA_DB):
        return None
    try:
        with open(FLOTA_DB, encoding="utf-8") as f:
            flota = json.load(f)
        return flota.get(veh_id.strip().upper())
    except Exception:
        return None


def _carpeta_soporte(vehiculo):
    p = SOPORTES_DIR / f"{datetime.now().strftime('%Y-%m-%d')}_{vehiculo.upper().replace(' ', '_')}"
    p.mkdir(parents=True, exist_ok=True)
    return p


def _guardar_soporte(archivo, vehiculo):
    folder = _carpeta_soporte(vehiculo)
    hora = datetime.now().strftime("%H%M%S")
    ext = archivo.name.split(".")[-1]
    fp = folder / f"soporte_{vehiculo.upper()}_{hora}.{ext}"
    fp.write_bytes(archivo.read())
    return str(fp)


def _cargar_pendientes():
    if not os.path.exists(PENDIENTES_DB):
        return {}
    try:
        with open(PENDIENTES_DB, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _guardar_pendientes(data):
    with open(PENDIENTES_DB, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _obtener_ot_pendiente(ot):
    return _cargar_pendientes().get(ot)


def _cerrar_ot_pendiente(ot):
    db = _cargar_pendientes()
    if ot in db:
        del db[ot]
        _guardar_pendientes(db)


def _listar_ots_pendientes():
    return sorted(_cargar_pendientes().keys())


def _actualizar_flota_tras_preventivo(vehiculo, rutina_ids, km_cierre):
    """Actualiza flota_vehiculos.json registrando las rutinas ejecutadas."""
    if not os.path.exists(FLOTA_DB):
        return
    try:
        with open(FLOTA_DB, "r", encoding="utf-8") as f:
            flota = json.load(f)
    except Exception:
        return

    vkey = vehiculo.upper()
    if vkey not in flota:
        return

    if "rutinas_ultimas" not in flota[vkey]:
        flota[vkey]["rutinas_ultimas"] = {}

    fecha_hoy = datetime.now().strftime("%Y-%m-%d")
    for rid in rutina_ids:
        flota[vkey]["rutinas_ultimas"][rid] = {"km": km_cierre, "fecha": fecha_hoy}

    # Actualizar km_actual si es mayor al registrado
    if km_cierre > flota[vkey].get("km_actual", 0):
        flota[vkey]["km_actual"] = km_cierre
        flota[vkey]["km_actualizado_en"] = fecha_hoy

    with open(FLOTA_DB, "w", encoding="utf-8") as f:
        json.dump(flota, f, ensure_ascii=False, indent=2)


def _save_excel_filas(filas):
    try:
        df_new = pd.DataFrame(filas, columns=COLUMNAS_EXCEL)
        if os.path.exists(EXCEL_FILE):
            try:
                df_old = pd.read_excel(EXCEL_FILE, engine="openpyxl")
                df_all = pd.concat([df_old, df_new], ignore_index=True)
            except PermissionError:
                return False, "El archivo Excel está abierto. Ciérrelo e intente de nuevo."
        else:
            df_all = df_new
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as w:
            df_all.to_excel(w, index=False, sheet_name="Mantenimiento")
            ws = w.sheets["Mantenimiento"]
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            hf = PatternFill("solid", fgColor="1A0D00")
            for cell in ws[1]:
                cell.fill = hf
                cell.font = Font(color="FF6B00", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            fa = PatternFill("solid", fgColor="161B22")
            fb = PatternFill("solid", fgColor="0D1117")
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
                for cell in row:
                    cell.fill = fa if i % 2 == 0 else fb
                    cell.alignment = Alignment(wrap_text=True)
            for col in ws.columns:
                ml = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 40)
        return True, "OK"
    except PermissionError:
        return False, "El archivo Excel está abierto."
    except Exception as e:
        return False, str(e)


def _volver_al_hub():
    for k in ["fase", "ot_actual", "intervenciones"]:
        st.session_state.pop(k, None)
    st.session_state["modulo"] = "hub"
    try:
        st.query_params.clear()
    except Exception:
        pass
    st.rerun()


# ═══════════════════════════════════════════════════════════════════
#  CSS
# ═══════════════════════════════════════════════════════════════════
CSS = """<style>
@import url('https://fonts.googleapis.com/css2?family=Exo+2:wght@700;800&family=Inter:wght@300;400;500;600&display=swap');
:root{--primary:#3FB950;--primary-dk:#2EA043;--bg:#0D1117;--surface:#161B22;--surface2:#21262D;
      --border:#30363D;--text:#E6EDF3;--muted:#8B949E;--green:#3FB950;--yellow:#D29922;
      --red:#F85149;--blue:#0A84FF;--orange:#FF6B00;--purple:#7C4DFF;--radius:12px;}
html,body,[data-testid="stAppViewContainer"]{background:var(--bg)!important;color:var(--text)!important;font-family:'Inter',sans-serif}
.app-header{background:linear-gradient(135deg,#0D1117,#0D1F0D,#0D1117);border-bottom:1px solid #1F3D1F;padding:1.3rem 2rem;margin:0 -1rem 2rem -1rem;display:flex;align-items:center;gap:1rem}
.app-header h1{font-family:'Exo 2',sans-serif;font-weight:800;font-size:1.5rem;color:var(--text);margin:0}
.app-header .sub{font-size:.78rem;color:var(--muted);margin:0}
.fase-badge{background:rgba(63,185,80,.15);color:var(--green);border:1px solid var(--green);padding:.25rem .9rem;border-radius:20px;font-size:.72rem;font-weight:700;text-transform:uppercase;letter-spacing:.05em}
.card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:1.5rem;margin-bottom:1.2rem}
.card-title{font-family:'Exo 2',sans-serif;font-size:.95rem;font-weight:700;color:var(--primary);margin-bottom:1rem;text-transform:uppercase;letter-spacing:.05em;border-bottom:1px solid var(--border);padding-bottom:.6rem}
.ot-loaded{background:linear-gradient(135deg,#0D1A00,#131F00);border:1.5px solid var(--green);border-radius:var(--radius);padding:1rem 1.5rem;margin:1rem 0}
.ot-label{font-size:.7rem;color:var(--muted);text-transform:uppercase;letter-spacing:.1em}
.ot-value{font-family:'Exo 2',sans-serif;font-size:1.3rem;font-weight:800;color:var(--green)}
.chip{display:inline-block;padding:2px 10px;border-radius:20px;font-size:.7rem;font-weight:700;text-transform:uppercase;letter-spacing:.05em;margin-right:.3rem}
.chip-ci{background:rgba(10,132,255,.15);color:var(--blue);border:1px solid rgba(10,132,255,.4)}
.chip-co{background:rgba(255,107,0,.15);color:var(--orange);border:1px solid rgba(255,107,0,.4)}
.chip-p {background:rgba(124,77,255,.15);color:var(--purple);border:1px solid rgba(124,77,255,.4)}
.chip-m {background:rgba(248,81,73,.15);color:var(--red);border:1px solid rgba(248,81,73,.4)}
.chip-crit-alta{background:rgba(248,81,73,.15);color:var(--red);border:1px solid var(--red);padding:2px 10px;border-radius:20px;font-size:.7rem;font-weight:700}
.chip-crit-media{background:rgba(210,153,34,.15);color:var(--yellow);border:1px solid var(--yellow);padding:2px 10px;border-radius:20px;font-size:.7rem;font-weight:700}
.chip-crit-baja{background:rgba(63,185,80,.12);color:var(--green);border:1px solid var(--green);padding:2px 10px;border-radius:20px;font-size:.7rem;font-weight:700}
.preventivo-info{background:rgba(124,77,255,.05);border-left:3px solid var(--purple);border-radius:6px;padding:.7rem 1rem;margin:0 0 1rem 0;font-size:.85rem}
.rutina-row{background:var(--surface2);border:1px solid var(--border);border-radius:6px;padding:.5rem .8rem;margin-bottom:.3rem;font-size:.82rem;color:var(--muted)}
.rutina-row strong{color:var(--text)}
.interv-header{display:grid;grid-template-columns:3fr 1fr 1.5fr .5fr;gap:.5rem;padding:.5rem .8rem;background:var(--surface2);border-radius:6px 6px 0 0;border:1px solid var(--border);border-bottom:none;font-size:.75rem;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.05em}
.interv-row{background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:.6rem .8rem;margin-bottom:.5rem}
.instruc-box{background:rgba(255,107,0,.07);border:1px dashed #FF6B00;border-radius:8px;padding:.6rem 1rem;margin-bottom:1rem;font-size:.8rem;color:var(--muted)}
.instruc-box strong{color:var(--orange)}
.costo-display{background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:.8rem 1.2rem;text-align:center}
.costo-label{font-size:.7rem;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;margin-bottom:.3rem}
.costo-valor{font-family:'Exo 2',sans-serif;font-size:1.55rem;font-weight:800;color:var(--primary);letter-spacing:.01em}
.costo-total{color:var(--green)!important}
.success-screen{text-align:center;padding:2.5rem 2rem;background:linear-gradient(135deg,#0D1F0D,#0D1117);border:1px solid var(--green);border-radius:var(--radius);margin:1.5rem 0}
.success-screen h2{font-family:'Exo 2',sans-serif;font-size:1.8rem;color:var(--green);margin:.4rem 0}
.origen-chip{display:inline-block;background:rgba(10,132,255,.15);color:var(--blue);border:1px solid rgba(10,132,255,.3);padding:2px 10px;border-radius:20px;font-size:.72rem;font-weight:600;margin-left:.4rem}
[data-testid="stTextInput"] input,[data-testid="stTextArea"] textarea,[data-testid="stNumberInput"] input{background:var(--surface2)!important;border:1px solid var(--border)!important;color:var(--text)!important;border-radius:8px!important;font-family:'Inter',sans-serif!important}
[data-testid="stSelectbox"]>div>div{background:var(--surface2)!important;border:1px solid var(--border)!important;border-radius:8px!important}
[data-testid="stRadio"]>div{gap:.4rem!important}
[data-testid="stRadio"] label{background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:.35rem .9rem!important;cursor:pointer;font-size:.85rem!important}
[data-testid="stRadio"] label:has(input:checked){border-color:var(--green)!important;background:rgba(63,185,80,.1)!important}
[data-testid="stButton"] button{background:var(--primary)!important;color:white!important;border:none!important;border-radius:8px!important;font-family:'Exo 2',sans-serif!important;font-weight:700!important;font-size:.9rem!important;padding:.55rem 1.2rem!important;text-transform:uppercase!important;letter-spacing:.05em!important}
[data-testid="stButton"] button:hover{background:var(--primary-dk)!important;transform:translateY(-1px);box-shadow:0 4px 16px rgba(63,185,80,.3)!important}
[data-testid="stButton"] button:disabled{background:var(--surface2)!important;color:var(--muted)!important;transform:none!important;box-shadow:none!important}
.btn-blue button{background:#005FCC!important}.btn-blue button:hover{background:#0A84FF!important}
.btn-gray button{background:var(--surface2)!important;color:var(--muted)!important;border:1px solid var(--border)!important}
.btn-gray button:hover{background:var(--border)!important;color:var(--text)!important;transform:none!important;box-shadow:none!important}
.btn-back button{background:var(--surface2)!important;color:var(--muted)!important;border:1px solid var(--border)!important;text-transform:none!important;font-size:.82rem!important;font-weight:500!important;padding:.4rem 1rem!important;letter-spacing:normal!important}
.btn-back button:hover{background:var(--border)!important;color:var(--text)!important;transform:none!important;box-shadow:none!important}
.char-count{font-size:.72rem;color:var(--muted);text-align:right;margin-top:-.3rem}
.char-count-warn{color:var(--yellow)}
.char-count-danger{color:var(--red)}
#MainMenu,footer,header{visibility:hidden}.block-container{padding-top:0!important}
</style>"""


# ═══════════════════════════════════════════════════════════════════
#  ESTADO
# ═══════════════════════════════════════════════════════════════════
def _init_state():
    defaults = {
        "fase":           "2",  # El módulo arranca directo en Fase 2 (cierre)
        "ot_actual":      None,
        "intervenciones": [{"desc": "", "cantidad": 1, "costo": 0.0}],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _render_back_button(key_suffix):
    if st.session_state.get("modulo") == "cierre_ot":
        st.markdown('<div class="btn-back">', unsafe_allow_html=True)
        if st.button("← Volver al Hub", key=f"btn_back_co_{key_suffix}"):
            _volver_al_hub()
        st.markdown('</div>', unsafe_allow_html=True)


def _header(badge_txt="Cierre Técnico"):
    st.markdown(f"""
    <div class="app-header">
      <div style="font-size:2rem">📋</div>
      <div>
        <h1>Cierre de Orden de Trabajo</h1>
        <p class="sub">Completar intervención técnica · Soporta OT-CI, OT-CO, OT-P, OT-M</p>
      </div>
      <div style="margin-left:auto"><span class="fase-badge">{badge_txt}</span></div>
    </div>""", unsafe_allow_html=True)


def _chip_tipo_ot(ot_num):
    """Retorna el chip HTML del tipo según el prefijo de la OT."""
    if ot_num.startswith("OT-CI-"):
        return '<span class="chip chip-ci">CORR. INSPECCIÓN</span>'
    elif ot_num.startswith("OT-CO-"):
        return '<span class="chip chip-co">CORR. OPERACIÓN</span>'
    elif ot_num.startswith("OT-P-"):
        return '<span class="chip chip-p">PREVENTIVO</span>'
    elif ot_num.startswith("OT-M-"):
        return '<span class="chip chip-m">CORR. MAYOR</span>'
    return ''


def _chip_criticidad(crit):
    if crit == "Alta":  return '<span class="chip-crit-alta">🔴 Alta</span>'
    if crit == "Media": return '<span class="chip-crit-media">🟡 Media</span>'
    if crit == "Baja":  return '<span class="chip-crit-baja">🟢 Baja</span>'
    return ''


# ═══════════════════════════════════════════════════════════════════
#  run()
# ═══════════════════════════════════════════════════════════════════
def run():
    _init_state()
    st.markdown(CSS, unsafe_allow_html=True)

    fase = st.session_state.get("fase", "2")

    # ══════════════════════════════════════════════
    #  CONFIRMACIÓN POST-CIERRE
    # ══════════════════════════════════════════════
    if fase == "confirmacion_f2":
        _render_back_button("conf")
        _header("OT Finalizada")
        ot = st.session_state.get("ot_actual", "")
        tipo_chip = _chip_tipo_ot(ot)

        st.markdown(f"""<div class="success-screen">
          <div style="font-size:4rem">✅</div>
          <h2>Orden de Trabajo Cerrada</h2>
          <p style="color:#8B949E;margin:.4rem 0">La hoja de vida fue actualizada. Estado: <strong style="color:var(--green)">Finalizada</strong>.</p>
          <div class="ot-loaded" style="display:inline-block;text-align:left;margin-top:.8rem">
            <div class="ot-label">OT Cerrada {tipo_chip}</div>
            <div class="ot-value">{ot}</div>
          </div>
        </div>""", unsafe_allow_html=True)

        col_a, col_b, col_c = st.columns(3)
        with col_a:
            if st.button("Cerrar otra OT", use_container_width=True, key="cf2_otra"):
                st.session_state["fase"] = "2"
                st.session_state["ot_actual"] = None
                st.session_state["intervenciones"] = [{"desc": "", "cantidad": 1, "costo": 0.0}]
                st.rerun()
        with col_b:
            st.markdown('<div class="btn-blue">', unsafe_allow_html=True)
            if st.button("Ir a Reporte de Fallas", use_container_width=True, key="cf2_rf"):
                st.session_state["modulo"] = "reporte_fallas"
                try:
                    st.query_params["mod"] = "reporte_fallas"
                except Exception:
                    pass
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        with col_c:
            st.markdown('<div class="btn-gray">', unsafe_allow_html=True)
            if st.button("Volver al Hub", use_container_width=True, key="cf2_hub"):
                _volver_al_hub()
            st.markdown('</div>', unsafe_allow_html=True)
        return

    # ══════════════════════════════════════════════
    #  FASE 2 — CIERRE TÉCNICO
    # ══════════════════════════════════════════════
    _render_back_button("f2")
    _header()
    st.markdown(f"<p style='color:#8B949E;font-size:.8rem;margin-bottom:1.5rem'>📅 {datetime.now().strftime('%A, %d de %B de %Y — %H:%M')}</p>",
                unsafe_allow_html=True)

    # ── Cargar OT ──────────────────────────────────
    st.markdown('<div class="card"><div class="card-title">🔍 Seleccionar Orden de Trabajo</div>',
                unsafe_allow_html=True)

    pendientes = _listar_ots_pendientes()
    ot_cargada = st.session_state.get("ot_actual")
    f1 = _obtener_ot_pendiente(ot_cargada) if ot_cargada else None

    if not f1:
        # ── FILTRO POR VEHÍCULO ─────────────────────────
        # Construir mapa de vehículos a partir de las OTs pendientes
        vehiculos_con_ot = sorted(set(
            (_obtener_ot_pendiente(o) or {}).get("Numero_Interno", "")
            for o in pendientes
        ) - {""})

        if pendientes and len(vehiculos_con_ot) > 1:
            col_filt, _ = st.columns([2, 2])
            with col_filt:
                opciones_filtro = ["Todos los vehículos"] + vehiculos_con_ot
                veh_filtro = st.selectbox(
                    "🔍 Filtrar por vehículo:",
                    options=opciones_filtro,
                    key="filtro_veh_ot",
                    help="Filtre las OTs pendientes por número interno"
                )
            if veh_filtro != "Todos los vehículos":
                pendientes_visibles = [
                    o for o in pendientes
                    if (_obtener_ot_pendiente(o) or {}).get("Numero_Interno", "") == veh_filtro
                ]
                # Mostrar info del vehículo
                info_veh = _info_vehiculo(veh_filtro)
                if info_veh:
                    marca = info_veh.get("marca", "")
                    modelo = info_veh.get("modelo", "")
                    km_act = info_veh.get("km_actual", 0)
                    detalles = []
                    if marca or modelo:
                        detalles.append(f"🚛 {marca} {modelo}".strip())
                    if km_act > 0:
                        detalles.append(f"{km_act:,} km".replace(",", "."))
                    if detalles:
                        st.caption(" · ".join(detalles))
            else:
                pendientes_visibles = pendientes
        else:
            pendientes_visibles = pendientes

        col_r1, col_r2 = st.columns([3, 1])
        with col_r1:
            if pendientes_visibles:
                # Mostrar con chip de tipo al lado
                def _fmt_opt(x):
                    if x.startswith("OT-CI-"): return f"{x}   [Insp]"
                    if x.startswith("OT-CO-"): return f"{x}   [Oper]"
                    if x.startswith("OT-P-"):  return f"{x}   [Prev]"
                    if x.startswith("OT-M-"):  return f"{x}   [May]"
                    return x
                opciones = ["-- Escribir manualmente --"] + pendientes_visibles
                sel = st.selectbox("Seleccionar OT pendiente:",
                                     options=opciones,
                                     format_func=lambda x: _fmt_opt(x) if x in pendientes_visibles else x,
                                     key="sel_rad")
                if sel == "-- Escribir manualmente --":
                    radicado_input = st.text_input("Radicado OT *",
                                                    placeholder="Ej: OT-CO-FRE-0042-250423-1430",
                                                    key="txt_rad")
                else:
                    radicado_input = sel
            elif pendientes:
                # Hay OTs pero el filtro las oculta
                st.info(f"No hay OTs pendientes para el vehículo filtrado.")
                radicado_input = ""
            else:
                radicado_input = st.text_input("Radicado OT *",
                                                placeholder="Ej: OT-CO-FRE-0042-250423-1430",
                                                key="txt_rad_solo")
        with col_r2:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            if st.button("Cargar OT", key="btn_cargar"):
                ot_in = radicado_input.strip() if isinstance(radicado_input, str) else ""
                if _obtener_ot_pendiente(ot_in):
                    st.session_state["ot_actual"] = ot_in
                    st.rerun()
                else:
                    st.error(f"No se encontró la OT '{ot_in}'.")

        if pendientes:
            # Desglose por tipo
            n_ci = sum(1 for x in pendientes if x.startswith("OT-CI-"))
            n_co = sum(1 for x in pendientes if x.startswith("OT-CO-"))
            n_p  = sum(1 for x in pendientes if x.startswith("OT-P-"))
            n_m  = sum(1 for x in pendientes if x.startswith("OT-M-"))
            desglose = []
            if n_ci: desglose.append(f"{n_ci} inspección")
            if n_co: desglose.append(f"{n_co} operación")
            if n_p:  desglose.append(f"{n_p} preventivo")
            if n_m:  desglose.append(f"{n_m} mayor")
            st.caption(f"📊 {len(pendientes)} OTs pendientes: " + " · ".join(desglose))
        else:
            st.info("No hay OTs pendientes. Reporte una falla o planifique un preventivo para generarlas.")
        st.markdown('</div>', unsafe_allow_html=True)
        return

    # ── OT Cargada — determinar tipo ──────────────
    tipo_mant    = f1.get("Tipo_Mantenimiento", "C")
    es_correct   = tipo_mant in TIPOS_CORRECTIVOS
    es_preventivo = tipo_mant == "P"
    origen       = f1.get("Origen", "")
    criticidad   = f1.get("Criticidad", "")
    origen_chip  = f'<span class="origen-chip">Origen: {origen.replace("_", " ")}</span>' if origen else ""
    crit_chip    = _chip_criticidad(criticidad) if criticidad else ""
    tipo_chip    = _chip_tipo_ot(ot_cargada)

    st.markdown(f"""<div class="ot-loaded">
      <div class="ot-label">OT cargada — completando cierre técnico {origen_chip}</div>
      <div class="ot-value">{ot_cargada} &nbsp; {tipo_chip}</div>
      <div style="margin-top:.5rem;font-size:.8rem;color:#8B949E">
        Vehículo: <strong style="color:#E6EDF3">{f1.get('Numero_Interno', '')}</strong> &nbsp;·&nbsp;
        Tipo: <strong style="color:#E6EDF3">{TIPOS_MANT.get(tipo_mant, '—')}</strong> &nbsp;·&nbsp;
        Sistema: <strong style="color:#E6EDF3">{f1.get('Sistema', '')}</strong>
        {('&nbsp;·&nbsp; Criticidad: ' + crit_chip) if crit_chip else ''}<br>
        {'Rutinas planificadas' if es_preventivo else 'Falla reportada'}:
        <strong style="color:{'var(--purple)' if es_preventivo else 'var(--red)'}">{f1.get('Modo_Falla', '—')}</strong> &nbsp;·&nbsp;
        Reporte: <strong style="color:#E6EDF3">{f1.get('Fecha_Registro_F1', '')}</strong>
      </div>
    </div>""", unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # ── PRE-CARGA DE INFORMACIÓN (solo preventivo) ──
    if es_preventivo and "Rutinas_IDs" in f1:
        st.markdown(f"""<div class="preventivo-info">
          🔧 <strong style="color:var(--purple)">OT Preventiva planificada</strong> — Rutinas del catálogo a ejecutar:<br>
          <div style="margin-top:.5rem;font-size:.8rem">
            {f1.get('Rutinas_Nombres', '')}
          </div>
          <div style="margin-top:.6rem;font-size:.78rem;color:var(--muted)">
            <strong>Costo estimado:</strong> {_fmt_moneda(f1.get('Costo_Estimado_Total', 0))} &nbsp;·&nbsp;
            <strong>Duración estimada:</strong> {f1.get('Duracion_Estimada_Horas', 0)} h &nbsp;·&nbsp;
            <strong>Responsable:</strong> {f1.get('Responsable_Planificado', '')}
          </div>
          {'<div style="margin-top:.6rem;font-size:.78rem;color:var(--muted)"><strong>Repuestos planificados:</strong> ' + f1.get('Repuestos_Planificados', '') + '</div>' if f1.get('Repuestos_Planificados') else ''}
        </div>""", unsafe_allow_html=True)

    # ── Diagnóstico ──
    st.markdown('<div class="card"><div class="card-title">🔎 Diagnóstico y Responsables</div>',
                unsafe_allow_html=True)

    causa_raiz = ""
    if es_correct:
        st.markdown(f"""<div style="background:rgba(248,81,73,.05);border:1px solid rgba(248,81,73,.2);border-radius:8px;padding:.6rem 1rem;margin-bottom:1rem;font-size:.78rem;color:#F85149">
          ⚠️ <strong>Causa raíz del modo de falla reportado</strong>
          <span style="color:#8B949E"> · máximo 100 caracteres · obligatorio para correctivos.</span>
        </div>""", unsafe_allow_html=True)
        causa_raiz = st.text_area(
            f"Causa raíz de: \"{f1.get('Modo_Falla', '')[:60]}...\" *",
            placeholder="Ej: Desgaste prematuro por falta de mantenimiento preventivo",
            height=80, max_chars=100, key="causa_raiz_mant"
        )
        ch_count = len(causa_raiz)
        ch_class = "char-count"
        if ch_count >= 90:   ch_class += " char-count-danger"
        elif ch_count >= 70: ch_class += " char-count-warn"
        st.markdown(f'<div class="{ch_class}">{ch_count} / 100 caracteres</div>',
                    unsafe_allow_html=True)
    else:
        st.info("ℹ️ Mantenimiento preventivo — no requiere causa raíz.")
        causa_raiz = "N/A - Preventivo"

    col_cr1, col_cr2 = st.columns(2)
    with col_cr1:
        # Pre-llenar con responsable planificado si viene de preventivo
        resp_default = f1.get("Responsable_Planificado", "") if es_preventivo else ""
        responsable = st.text_input("Responsable técnico que ejecutó *",
                                     value=resp_default,
                                     placeholder="Ej: Mecánico Juan Torres",
                                     key="resp_tec")
    with col_cr2:
        prov_default = f1.get("Proveedor_Planificado", "") if es_preventivo else ""
        proveedor = st.text_input("Proveedor / Taller *",
                                    value=prov_default,
                                    placeholder="Ej: Taller El Rápido",
                                    key="proveedor_f2")
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Intervenciones ──
    st.markdown('<div class="card"><div class="card-title">🔩 Intervenciones Realizadas</div>',
                unsafe_allow_html=True)
    st.markdown("""<div class="instruc-box">
      <strong>📌 Formato:</strong>&nbsp;<strong>VERBO + COMPONENTE + MARCA + REFERENCIA</strong><br>
      <span style="color:#8B949E">Ej: CAMBIO filtro aceite Fleetguard LF3349 &nbsp;·&nbsp; REEMPLAZO pastillas Brembo P50-012</span>
    </div>""", unsafe_allow_html=True)

    st.markdown("""<div class="interv-header">
      <span>Descripción</span><span style="text-align:center">Cant.</span>
      <span style="text-align:center">Costo Repuestos</span><span></span>
    </div>""", unsafe_allow_html=True)

    intervenciones = st.session_state["intervenciones"]
    to_del = []
    for idx, interv in enumerate(intervenciones):
        st.markdown('<div class="interv-row">', unsafe_allow_html=True)
        ci1, ci2, ci3, ci4 = st.columns([3, 1, 1.5, 0.5])
        with ci1:
            d = st.text_input(f"_d{idx}", value=interv["desc"],
                               placeholder="Ej: CAMBIO aceite Valvoline 20W50",
                               key=f"desc_{idx}", label_visibility="collapsed")
            intervenciones[idx]["desc"] = d
        with ci2:
            c = st.number_input(f"_c{idx}", min_value=1, step=1,
                                 value=int(interv["cantidad"]),
                                 key=f"cant_{idx}", label_visibility="collapsed")
            intervenciones[idx]["cantidad"] = c
        with ci3:
            k = st.number_input(f"_k{idx}", min_value=0.0, step=1000.0,
                                 value=float(interv["costo"]),
                                 key=f"costo_{idx}", label_visibility="collapsed",
                                 format="%.0f")
            intervenciones[idx]["costo"] = k
            if k > 0:
                st.markdown(f"<div style='font-size:.95rem;font-weight:700;color:#FF6B00;margin-top:-.3rem'>"
                            f"{_fmt_moneda(k)}</div>", unsafe_allow_html=True)
        with ci4:
            if len(intervenciones) > 1:
                if st.button("✕", key=f"del_{idx}"):
                    to_del.append(idx)
        st.markdown('</div>', unsafe_allow_html=True)

    for idx in sorted(to_del, reverse=True):
        intervenciones.pop(idx)
    st.session_state["intervenciones"] = intervenciones

    st.markdown('<div class="btn-blue">', unsafe_allow_html=True)
    if st.button("➕ Agregar intervención", key="btn_add"):
        st.session_state["intervenciones"].append({"desc": "", "cantidad": 1, "costo": 0.0})
        st.rerun()
    st.markdown('</div></div>', unsafe_allow_html=True)

    # ── Costos ──
    st.markdown('<div class="card"><div class="card-title">💰 Costos Reales</div>',
                unsafe_allow_html=True)
    total_rep = sum(i["costo"] for i in st.session_state["intervenciones"])

    if es_preventivo:
        col_mo, col_tr, col_tot, col_est = st.columns(4)
    else:
        col_mo, col_tr, col_tot = st.columns(3)

    with col_mo:
        mano_obra = st.number_input("Costo mano de obra *", min_value=0.0,
                                      step=10000.0, value=0.0, key="mano_obra",
                                      format="%.0f")
        if mano_obra > 0:
            st.markdown(f"<div style='font-size:.95rem;font-weight:700;color:#FF6B00;margin-top:-.3rem'>"
                        f"{_fmt_moneda(mano_obra)}</div>", unsafe_allow_html=True)
    costo_total = total_rep + mano_obra
    with col_tr:
        st.markdown(f'<div class="costo-display"><div class="costo-label">Total Repuestos</div>'
                    f'<div class="costo-valor">{_fmt_moneda(total_rep)}</div></div>',
                    unsafe_allow_html=True)
    with col_tot:
        st.markdown(f'<div class="costo-display"><div class="costo-label">Costo Total Real</div>'
                    f'<div class="costo-valor costo-total">{_fmt_moneda(costo_total)}</div></div>',
                    unsafe_allow_html=True)

    if es_preventivo:
        costo_est = f1.get("Costo_Estimado_Total", 0)
        diff = costo_total - costo_est
        diff_color = "#F85149" if diff > 0 else ("#3FB950" if diff < 0 else "#8B949E")
        diff_sym = "+" if diff > 0 else ""
        with col_est:
            st.markdown(f'<div class="costo-display"><div class="costo-label">Diff vs Estimado</div>'
                        f'<div class="costo-valor" style="color:{diff_color}">{diff_sym}{_fmt_moneda(abs(diff)) if diff != 0 else "$ 0"}</div></div>',
                        unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Tiempos ──
    st.markdown('<div class="card"><div class="card-title">⏱️ Tiempos de Cierre</div>',
                unsafe_allow_html=True)
    col_tc1, col_tc2 = st.columns(2)
    with col_tc1:
        fecha_cierre = st.date_input("Fecha finalización OT *",
                                      value=datetime.now().date(), key="fecha_cierre")
    with col_tc2:
        hora_cierre = st.time_input("Hora finalización *",
                                      value=datetime.now().time(), key="hora_cierre")

    st.markdown("<div style='font-size:.85rem;font-weight:600;color:#8B949E;margin:.8rem 0 .4rem'>"
                "Tiempo de consecución del repuesto</div>", unsafe_allow_html=True)
    st.caption("Indique el tiempo total que tomó conseguir los repuestos. Use días para esperas largas (semanas) y horas para complementar.")
    col_tr1, col_tr2 = st.columns(2)
    with col_tr1:
        tiempo_rep_dias = st.number_input("Días", min_value=0, step=1, value=0,
                                            key="tiempo_rep_dias", format="%d")
    with col_tr2:
        tiempo_rep_horas = st.number_input("Horas (adicionales a los días)",
                                             min_value=0.0, step=0.5, max_value=23.5,
                                             value=0.0, key="tiempo_rep", format="%.1f")
    # Tiempo total expresado en horas (para almacenar)
    tiempo_rep_total_horas = tiempo_rep_dias * 24 + tiempo_rep_horas
    if tiempo_rep_total_horas > 0:
        if tiempo_rep_dias > 0 and tiempo_rep_horas > 0:
            txt = f"{tiempo_rep_dias} día{'s' if tiempo_rep_dias != 1 else ''} y {tiempo_rep_horas:.1f} h"
        elif tiempo_rep_dias > 0:
            txt = f"{tiempo_rep_dias} día{'s' if tiempo_rep_dias != 1 else ''}"
        else:
            txt = f"{tiempo_rep_horas:.1f} h"
        st.markdown(f"<div style='font-size:.85rem;color:var(--blue);font-weight:600;margin-top:.3rem'>"
                    f"⏰ Total: {txt}  ({tiempo_rep_total_horas:.1f} h equivalentes)</div>",
                    unsafe_allow_html=True)

    # Kilometraje al cierre (importante para preventivos — alimenta la flota)
    km_cierre = st.number_input("Kilometraje al cierre (actual del vehículo)",
                                 min_value=0, step=100,
                                 value=int(f1.get("Kilometraje", 0)),
                                 key="km_cierre", format="%d")
    if km_cierre > 0:
        st.markdown(f"<div style='font-size:.85rem;color:var(--primary);font-weight:600;margin-top:-.4rem'>"
                    f"{km_cierre:,} km".replace(",", ".") + "</div>", unsafe_allow_html=True)
    if es_preventivo:
        st.caption("🔄 Este kilometraje actualiza el registro de rutinas ejecutadas en la flota.")
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Soporte ──
    st.markdown('<div class="card"><div class="card-title">📎 Soporte / Factura</div>',
                unsafe_allow_html=True)
    st.caption("Adjuntar factura, remisión o soporte. PDF, Excel, Word, JPG, PNG.")
    soportes_f2 = st.file_uploader("Adjuntar soporte",
                                    type=["pdf", "xlsx", "xls", "docx", "doc", "jpg", "jpeg", "png"],
                                    accept_multiple_files=True, key="soportes_f2")
    if soportes_f2:
        for sf in soportes_f2:
            st.markdown(f"📄 **{sf.name}**")
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Validación ──
    interv_validas = [i for i in st.session_state["intervenciones"] if i["desc"].strip()]
    causa_ok = (not es_correct) or bool(causa_raiz.strip() and causa_raiz != "N/A - Preventivo")
    f2_ok = (causa_ok and bool(responsable.strip())
             and bool(proveedor.strip()) and len(interv_validas) > 0)

    falt = []
    if es_correct and not (causa_raiz.strip() and causa_raiz != "N/A - Preventivo"):
        falt.append("**Causa raíz**")
    if not responsable.strip(): falt.append("**Responsable técnico**")
    if not proveedor.strip():   falt.append("**Proveedor**")
    if not interv_validas:      falt.append("**Al menos una intervención**")
    if falt:
        st.warning(f"Campos pendientes: {' · '.join(falt)}")

    col_g2, col_v2 = st.columns([3, 1])
    with col_g2:
        guardar_f2 = st.button("✅ CERRAR ORDEN DE TRABAJO",
                                disabled=not f2_ok, use_container_width=True,
                                key="btn_guardar_f2")
    with col_v2:
        st.markdown('<div class="btn-gray">', unsafe_allow_html=True)
        if st.button("↩ Elegir otra OT", use_container_width=True, key="btn_cambiar",
                     help="Descarta la OT cargada y regresa al selector para cargar una OT diferente."):
            st.session_state["ot_actual"] = None
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    if guardar_f2 and f2_ok:
        soporte_paths = []
        if soportes_f2:
            for sf in soportes_f2:
                sf.seek(0)
                soporte_paths.append(_guardar_soporte(sf, f1.get("Numero_Interno", "")))

        filas = []
        for interv in interv_validas:
            filas.append({
                "Fecha_Inicio_Inactividad": f1.get("Fecha_Inicio_Inactividad", ""),
                "Fecha_Registro_F1":        f1.get("Fecha_Registro_F1", ""),
                "Fecha_Cierre_F2":          f"{fecha_cierre} {hora_cierre}",
                "OT":                       ot_cargada,
                "Numero_Interno":           f1.get("Numero_Interno", ""),
                "Tipo_Mantenimiento":       tipo_mant,
                "Estado_OT":                "F",
                "Criticidad":               f1.get("Criticidad", ""),
                "Kilometraje":              km_cierre,
                "Sistema":                  f1.get("Sistema", ""),
                "Modo_Falla":               f1.get("Modo_Falla", ""),
                "Causa_Raiz":               causa_raiz.strip() if es_correct else "N/A - Preventivo",
                "Descripcion":              interv["desc"].strip(),
                "Cantidad":                 int(interv["cantidad"]),
                "Costo_Repuestos":          float(interv["costo"]),
                "Costo_Mano_Obra":          float(mano_obra),
                "Costo_Total_OT":           float(costo_total),
                "Costo_Estimado_Total":     float(f1.get("Costo_Estimado_Total", 0)) if es_preventivo else "",
                "Proveedor":                proveedor.strip(),
                "Tiempo_Repuesto_Dias":     int(tiempo_rep_dias) if tiempo_rep_dias > 0 else "",
                "Tiempo_Repuesto_Horas":    float(tiempo_rep_total_horas) if tiempo_rep_total_horas > 0 else "",
                "Responsable_Tecnico":      responsable.strip(),
                "Conductor":                f1.get("Conductor", ""),
                "Soporte":                  " | ".join(soporte_paths),
                "Origen":                   f1.get("Origen", "Reporte_Directo"),
            })

        ok, msg = _save_excel_filas(filas)
        if ok:
            # Actualizar flota si era preventivo
            if es_preventivo and f1.get("Rutinas_IDs"):
                rutina_ids = [x.strip() for x in f1.get("Rutinas_IDs", "").split(",") if x.strip()]
                _actualizar_flota_tras_preventivo(f1.get("Numero_Interno", ""),
                                                    rutina_ids, km_cierre)

            _cerrar_ot_pendiente(ot_cargada)
            st.session_state["fase"] = "confirmacion_f2"
            st.rerun()
        else:
            st.error(f"Error: {msg}")


# ═══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    try:
        st.set_page_config(page_title="Cierre de OT",
                           page_icon="📋", layout="wide",
                           initial_sidebar_state="collapsed")
    except Exception:
        pass
    run()